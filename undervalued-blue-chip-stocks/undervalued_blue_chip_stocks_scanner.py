# -*- coding: utf-8 -*-
"""
screener_from_details_cache.py

(인터넷 연결 불필요) build_details_cache.py가 만든 details_cache_{source}.csv/.xlsx
하나만으로 4개 프로파일 결과를 Excel로 출력.

개선사항:
1. 버핏 스타일에 더 적합한 점수 체계
2. 섹터별 차별화된 평가
3. 현실적인 필터링 조건
4. 더 다양한 재무 지표 반영
5. 향상된 조건부 서식 (색상으로 가독성 향상)
"""

import os, math, time, random, warnings, openpyxl
import pandas as pd, numpy as np
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

warnings.filterwarnings("ignore", category=RuntimeWarning)


# 스타일 상수 정의
class ExcelStyles:
    """엑셀 스타일 정의 클래스"""

    # 색상 정의
    LIGHT_BLUE = "E6F3FF"
    LIGHT_GRAY = "F5F5F5"
    LIGHT_GREEN = "F0F8F0"
    LIGHT_YELLOW = "FFFDE6"
    HEADER_BLUE = "4F81BD"
    HEADER_FONT_COLOR = "FFFFFF"

    # 조건부 서식 색상
    GREEN = "009000"  # 좋은 수치
    LIGHT_GREEN = "90EE90"  # 양호한 수치
    YELLOW = "FFFF00"  # 주의 needed
    ORANGE = "FFA500"  # 경고
    RED = "FF0000"  # 위험

    # 채우기 패턴
    LIGHT_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    DARK_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    HEADER_FILL = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")

    # 폰트
    HEADER_FONT = Font(name='Calibri', size=11, bold=True, color=HEADER_FONT_COLOR)
    NORMAL_FONT = Font(name='Calibri', size=10)
    BOLD_FONT = Font(name='Calibri', size=10, bold=True)

    # 정렬
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
    LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
    RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')

    # 테두리
    THIN_BORDER = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # 숫자 포맷
    FORMAT_PERCENT = '0.00%'
    FORMAT_PERCENT_1 = '0.0%'
    FORMAT_CURRENCY = '#,##0.00'
    FORMAT_CURRENCY_INT = '#,##0'
    FORMAT_NUMBER_2 = '0.00'
    FORMAT_NUMBER_1 = '0.0'
    FORMAT_NUMBER_INT = '0'
    FORMAT_LARGE_NUMBER = '#,##0.00,, "B"'  # 10억 단위


def apply_number_formatting(worksheet, df, start_row=2):
    """
    컬럼별 숫자 포맷팅 적용
    """
    try:
        # 컬럼별 포맷 매핑
        format_mapping = {
            # 백분율 포맷 (2자리)
            'Discount_Pct': ExcelStyles.FORMAT_PERCENT,
            'DivYield': ExcelStyles.FORMAT_PERCENT,
            'ROE(info)': ExcelStyles.FORMAT_PERCENT,
            'ROE_5Y_Avg': ExcelStyles.FORMAT_PERCENT,  # ROE_5Y_Avg 추가
            'RevYoY': ExcelStyles.FORMAT_PERCENT,
            'OpMarginTTM': ExcelStyles.FORMAT_PERCENT,
            'OperatingMargins(info)': ExcelStyles.FORMAT_PERCENT,
            'FCF_Yield': ExcelStyles.FORMAT_PERCENT,
            'PayoutRatio': ExcelStyles.FORMAT_PERCENT,
            'ATR_PCT': ExcelStyles.FORMAT_PERCENT,
            'RET5': ExcelStyles.FORMAT_PERCENT,
            'RET20': ExcelStyles.FORMAT_PERCENT,

            # 통화 포맷
            'Price': ExcelStyles.FORMAT_CURRENCY,
            'FairValue_Composite': ExcelStyles.FORMAT_CURRENCY,
            'FairValue_DCF': ExcelStyles.FORMAT_CURRENCY,
            'FairValue_Relative': ExcelStyles.FORMAT_CURRENCY,
            'FairValue_DDM': ExcelStyles.FORMAT_CURRENCY,
            'FairValue_Graham': ExcelStyles.FORMAT_CURRENCY,
            'DollarVol($M)': ExcelStyles.FORMAT_CURRENCY,

            # 10억 단위 큰 숫자
            'MktCap($B)': ExcelStyles.FORMAT_LARGE_NUMBER,

            # 소수점 2자리 숫자
            'PE': ExcelStyles.FORMAT_NUMBER_2,
            'PEG': ExcelStyles.FORMAT_NUMBER_2,
            'PB': ExcelStyles.FORMAT_NUMBER_2,
            'EV_EBITDA': ExcelStyles.FORMAT_NUMBER_2,
            'P_FFO': ExcelStyles.FORMAT_NUMBER_2,
            'Debt_to_Equity': ExcelStyles.FORMAT_NUMBER_2,
            'RVOL': ExcelStyles.FORMAT_NUMBER_2,
            'SMA20': ExcelStyles.FORMAT_NUMBER_2,
            'SMA50': ExcelStyles.FORMAT_NUMBER_2,

            # 소수점 1자리 (점수들)
            'GrowthScore': ExcelStyles.FORMAT_NUMBER_1,
            'QualityScore': ExcelStyles.FORMAT_NUMBER_1,
            'ValueScore': ExcelStyles.FORMAT_NUMBER_1,
            'CatalystScore': ExcelStyles.FORMAT_NUMBER_1,
            'TotalScore': ExcelStyles.FORMAT_NUMBER_1,
            'ValuationAdjustedScore': ExcelStyles.FORMAT_NUMBER_1,
            'ModernBuffettScore': ExcelStyles.FORMAT_NUMBER_1,
            'TotalScore_Modern': ExcelStyles.FORMAT_NUMBER_1,
            'MomentumScore': ExcelStyles.FORMAT_NUMBER_1,
            'TrendScore': ExcelStyles.FORMAT_NUMBER_1,
            'LiquidityScore': ExcelStyles.FORMAT_NUMBER_1,
            'VolatilityScore': ExcelStyles.FORMAT_NUMBER_1,
        }

        # 컬럼 인덱스 찾기
        col_mapping = {col: idx + 1 for idx, col in enumerate(df.columns)}

        # 각 컬럼에 포맷 적용
        for col_name, format_str in format_mapping.items():
            if col_name in col_mapping:
                col_letter = get_column_letter(col_mapping[col_name])

                # 해당 컬럼의 모든 셀에 포맷 적용
                for row in range(start_row, len(df) + start_row):
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.number_format = format_str

        print("   ✅ 숫자 포맷팅 적용 완료")

    except Exception as e:
        print(f"   ⚠️ 숫자 포맷팅 적용 중 오류: {e}")


def apply_enhanced_conditional_formatting(worksheet, df, sheet_name, start_row=2):
    """
    향상된 조건부 서식 적용 (색상으로 가독성 향상)
    """
    try:
        # 컬럼 인덱스 찾기
        col_mapping = {col: idx + 1 for idx, col in enumerate(df.columns)}

        # 프로파일별 조건부 서식 적용
        if any(profile in sheet_name.lower() for profile in ['buffett', 'modern']):
            apply_buffett_conditional_formatting(worksheet, df, col_mapping, start_row)
        elif any(profile in sheet_name.lower() for profile in ['swing', 'daytrade']):
            apply_trading_conditional_formatting(worksheet, df, col_mapping, start_row)

        print(f"   ✅ {sheet_name} 조건부 서식 적용 완료")

    except Exception as e:
        print(f"   ⚠️ 조건부 서식 적용 중 오류: {e}")


def apply_buffett_conditional_formatting(worksheet, df, col_mapping, start_row):
    """
    버핏 스타일 조건부 서식 적용
    """
    # 할인율 (Discount_Pct)
    if 'Discount_Pct' in col_mapping:
        col_letter = get_column_letter(col_mapping['Discount_Pct'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 8% 이상 할인 (매우 좋음)
        green_rule = CellIsRule(operator='greaterThan', formula=['0.08'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 0-8% 할인 (주의)
        orange_rule = CellIsRule(operator='between', formula=['0', '0.08'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 프리미엄 (위험)
        red_rule = CellIsRule(operator='lessThan', formula=['0'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # ROE (ROE(info))
    if 'ROE(info)' in col_mapping:
        col_letter = get_column_letter(col_mapping['ROE(info)'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 15% 이상 (우량)
        green_rule = CellIsRule(operator='greaterThan', formula=['0.15'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 10-15% (보통)
        orange_rule = CellIsRule(operator='between', formula=['0.10', '0.15'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 10% 미만 (위험)
        red_rule = CellIsRule(operator='lessThan', formula=['0.10'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # PER (PE)
    if 'PE' in col_mapping:
        col_letter = get_column_letter(col_mapping['PE'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 15배 이하 (저PER)
        green_rule = CellIsRule(operator='lessThan', formula=['15'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 15-25배 (보통)
        orange_rule = CellIsRule(operator='between', formula=['15', '25'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 25배 초과 (고PER)
        red_rule = CellIsRule(operator='greaterThan', formula=['25'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # PBR (PB)
    if 'PB' in col_mapping:
        col_letter = get_column_letter(col_mapping['PB'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 1.5배 이하 (저PBR)
        green_rule = CellIsRule(operator='lessThan', formula=['1.5'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 1.5-3배 (보통)
        orange_rule = CellIsRule(operator='between', formula=['1.5', '3.0'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 3배 초과 (고PBR)
        red_rule = CellIsRule(operator='greaterThan', formula=['3.0'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # 부채비율 (Debt_to_Equity)
    if 'Debt_to_Equity' in col_mapping:
        col_letter = get_column_letter(col_mapping['Debt_to_Equity'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 0.5 이하 (매우 건전)
        green_rule = CellIsRule(operator='lessThan', formula=['0.5'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 0.5-1.5 (보통)
        orange_rule = CellIsRule(operator='between', formula=['0.5', '1.5'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 1.5 초과 (위험)
        red_rule = CellIsRule(operator='greaterThan', formula=['1.5'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # 영업이익률 (OpMarginTTM)
    if 'OpMarginTTM' in col_mapping:
        col_letter = get_column_letter(col_mapping['OpMarginTTM'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 15% 이상 (고효율)
        green_rule = CellIsRule(operator='greaterThan', formula=['0.15'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 8-15% (보통)
        orange_rule = CellIsRule(operator='between', formula=['0.08', '0.15'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 8% 미만 (저효율)
        red_rule = CellIsRule(operator='lessThan', formula=['0.08'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # 매출성장률 (RevYoY)
    if 'RevYoY' in col_mapping:
        col_letter = get_column_letter(col_mapping['RevYoY'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 8% 이상 (강한성장)
        green_rule = CellIsRule(operator='greaterThan', formula=['0.08'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 2-8% (보통성장)
        orange_rule = CellIsRule(operator='between', formula=['0.02', '0.08'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 2% 미만 (낮은성장)
        red_rule = CellIsRule(operator='lessThan', formula=['0.02'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # 배당수익률 (DivYield)
    if 'DivYield' in col_mapping:
        col_letter = get_column_letter(col_mapping['DivYield'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 2-6% (적정)
        green_rule = CellIsRule(operator='between', formula=['0.02', '0.06'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 0-2% 또는 6-8% (주의)
        orange_rule1 = CellIsRule(operator='between', formula=['0', '0.02'],
                                  font=Font(color=ExcelStyles.ORANGE, bold=True))
        orange_rule2 = CellIsRule(operator='between', formula=['0.06', '0.08'],
                                  font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 8% 초과 (위험) 또는 0% (배당없음)
        red_rule = CellIsRule(operator='greaterThan', formula=['0.08'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule1)
        worksheet.conditional_formatting.add(range_str, orange_rule2)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # 배당성향 (PayoutRatio)
    if 'PayoutRatio' in col_mapping:
        col_letter = get_column_letter(col_mapping['PayoutRatio'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 30-60% (적정)
        green_rule = CellIsRule(operator='between', formula=['0.30', '0.60'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 60-80% (주의)
        orange_rule = CellIsRule(operator='between', formula=['0.60', '0.80'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 80% 초과 (위험)
        red_rule = CellIsRule(operator='greaterThan', formula=['0.80'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # TotalScore (종합점수)
    if 'TotalScore' in col_mapping:
        col_letter = get_column_letter(col_mapping['TotalScore'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 70점 이상 (최우량)
        green_rule = CellIsRule(operator='greaterThan', formula=['70'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 60-70점 (우량)
        orange_rule = CellIsRule(operator='between', formula=['60', '70'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 60점 미만 (일반)
        red_rule = CellIsRule(operator='lessThan', formula=['60'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # EV/EBITDA
    if 'EV_EBITDA' in col_mapping:
        col_letter = get_column_letter(col_mapping['EV_EBITDA'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 5-12배 (좋음)
        green_rule = CellIsRule(operator='between', formula=['5', '12'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 12-18배 (보통)
        orange_rule = CellIsRule(operator='between', formula=['12', '18'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 18배 이상 (고평가)
        red_rule = CellIsRule(operator='greaterThan', formula=['18'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # PEG
    if 'PEG' in col_mapping:
        col_letter = get_column_letter(col_mapping['PEG'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 0.5-1.0 (매우 좋음)
        green_rule = CellIsRule(operator='between', formula=['0.5', '1.0'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 1.0-1.5 (보통)
        orange_rule = CellIsRule(operator='between', formula=['1.0', '1.5'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 1.5 이상 (고평가)
        red_rule = CellIsRule(operator='greaterThan', formula=['1.5'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # FCF_Yield (자유현금흐름 수익률)
    if 'FCF_Yield' in col_mapping:
        col_letter = get_column_letter(col_mapping['FCF_Yield'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 5% 이상 (우량)
        green_rule = CellIsRule(operator='greaterThan', formula=['0.05'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 2-5% (보통)
        orange_rule = CellIsRule(operator='between', formula=['0.02', '0.05'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 2% 미만 (약함)
        red_rule = CellIsRule(operator='lessThan', formula=['0.02'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # ROE_5Y_Avg (5년 평균 ROE) - 더 상세한 조건부 서식
    if 'ROE_5Y_Avg' in col_mapping:
        col_letter = get_column_letter(col_mapping['ROE_5Y_Avg'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 진한 초록색: 20% 이상 (탁월)
        dark_green_rule = CellIsRule(operator='greaterThan', formula=['0.20'],
                                     font=Font(color='006400', bold=True))  # 진한 초록색
        # 초록색: 15-20% (우량)
        green_rule = CellIsRule(operator='between', formula=['0.15', '0.20'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 10-15% (보통) - 굵게
        orange_rule = CellIsRule(operator='between', formula=['0.10', '0.15'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 10% 미만 (위험)
        red_rule = CellIsRule(operator='lessThan', formula=['0.10'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, dark_green_rule)
        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # ATR_PCT (평균 실제 범위) - 버핏에서는 낮은 변동성 선호
    if 'ATR_PCT' in col_mapping:
        col_letter = get_column_letter(col_mapping['ATR_PCT'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 1-5% (안정적)
        green_rule = CellIsRule(operator='between', formula=['0.01', '0.05'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 5-10% (변동성 있음)
        orange_rule = CellIsRule(operator='between', formula=['0.05', '0.10'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 10% 이상 (고변동성)
        red_rule = CellIsRule(operator='greaterThan', formula=['0.10'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # RVOL (상대 거래량) - 버핏에서는 적정 거래량 선호
    if 'RVOL' in col_mapping:
        col_letter = get_column_letter(col_mapping['RVOL'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 0.8-2.0 (적정)
        green_rule = CellIsRule(operator='between', formula=['0.8', '2.0'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 2.0-5.0 (과열 가능성)
        orange_rule = CellIsRule(operator='between', formula=['2.0', '5.0'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 5.0 이상 (매우 과열) 또는 0.8 미만 (관심낮음)
        red_rule1 = CellIsRule(operator='greaterThan', formula=['5.0'],
                               font=Font(color=ExcelStyles.RED, bold=True))
        red_rule2 = CellIsRule(operator='lessThan', formula=['0.8'],
                               font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule1)
        worksheet.conditional_formatting.add(range_str, red_rule2)

    # RET5 (5일 수익률) - 버핏에서는 안정적 수익률 선호
    if 'RET5' in col_mapping:
        col_letter = get_column_letter(col_mapping['RET5'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: -5%~+5% (안정적)
        green_rule = CellIsRule(operator='between', formula=['-0.05', '0.05'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: ±5-10% (변동성 있음)
        orange_rule1 = CellIsRule(operator='between', formula=['0.05', '0.10'],
                                  font=Font(color=ExcelStyles.ORANGE, bold=True))
        orange_rule2 = CellIsRule(operator='between', formula=['-0.10', '-0.05'],
                                  font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: ±10% 이상 (고변동성)
        red_rule1 = CellIsRule(operator='greaterThan', formula=['0.10'],
                               font=Font(color=ExcelStyles.RED, bold=True))
        red_rule2 = CellIsRule(operator='lessThan', formula=['-0.10'],
                               font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule1)
        worksheet.conditional_formatting.add(range_str, orange_rule2)
        worksheet.conditional_formatting.add(range_str, red_rule1)
        worksheet.conditional_formatting.add(range_str, red_rule2)

    # RET20 (20일 수익률) - 버핏에서는 안정적 수익률 선호
    if 'RET20' in col_mapping:
        col_letter = get_column_letter(col_mapping['RET20'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: -10%~+15% (안정적)
        green_rule = CellIsRule(operator='between', formula=['-0.10', '0.15'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: ±15-25% (변동성 있음)
        orange_rule1 = CellIsRule(operator='between', formula=['0.15', '0.25'],
                                  font=Font(color=ExcelStyles.ORANGE, bold=True))
        orange_rule2 = CellIsRule(operator='between', formula=['-0.25', '-0.10'],
                                  font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: ±25% 이상 (고변동성)
        red_rule1 = CellIsRule(operator='greaterThan', formula=['0.25'],
                               font=Font(color=ExcelStyles.RED, bold=True))
        red_rule2 = CellIsRule(operator='lessThan', formula=['-0.25'],
                               font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule1)
        worksheet.conditional_formatting.add(range_str, orange_rule2)
        worksheet.conditional_formatting.add(range_str, red_rule1)
        worksheet.conditional_formatting.add(range_str, red_rule2)

    # SMA20 (20일 이동평균) - 주가와의 관계를 나타내는 지표
    if 'SMA20' in col_mapping and 'Price' in col_mapping:
        # 주가 대비 SMA20 비율 계산 (간접적 방법)
        # 실제로는 주가와 SMA20의 관계를 보여주는 별도 컬럼이 필요하지만,
        # 여기서는 SMA20 값 자체에 대한 조건부 서식은 적용하지 않음
        # 대신 트레이딩 스타일에서 주가와의 관계를 다룸
        pass

    # SMA50 (50일 이동평균) - 주가와의 관계를 나타내는 지표
    if 'SMA50' in col_mapping and 'Price' in col_mapping:
        # SMA20과 동일한 이유로 조건부 서식 적용하지 않음
        pass

    # OperatingMargins(info) (영업이익률)
    if 'OperatingMargins(info)' in col_mapping:
        col_letter = get_column_letter(col_mapping['OperatingMargins(info)'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 12% 이상 (우량)
        green_rule = CellIsRule(operator='greaterThan', formula=['0.12'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 8-12% (보통)
        orange_rule = CellIsRule(operator='between', formula=['0.08', '0.12'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 8% 미만 (저효율)
        red_rule = CellIsRule(operator='lessThan', formula=['0.08'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # _OpMarginUse (사용된 영업이익률)
    if '_OpMarginUse' in col_mapping:
        col_letter = get_column_letter(col_mapping['_OpMarginUse'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # OperatingMargins(info)와 동일한 기준 적용
        green_rule = CellIsRule(operator='greaterThan', formula=['0.12'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        orange_rule = CellIsRule(operator='between', formula=['0.08', '0.12'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        red_rule = CellIsRule(operator='lessThan', formula=['0.08'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    score_columns = ['GrowthScore', 'QualityScore', 'ValueScore', 'CatalystScore', 'TotalScore']
    for score_col in score_columns:
        if score_col in col_mapping:
            col_letter = get_column_letter(col_mapping[score_col])
            range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

            # 초록색: 0.7점 이상 (강함)
            green_rule = CellIsRule(operator='greaterThan', formula=['0.7'],
                                    font=Font(color=ExcelStyles.GREEN, bold=True))
            # 주황색: 0.6-0.7점 (보통)
            orange_rule = CellIsRule(operator='between', formula=['0.6', '0.7'],
                                     font=Font(color=ExcelStyles.ORANGE, bold=True))
            # 빨간색: 0.6점 미만 (약함)
            red_rule = CellIsRule(operator='lessThan', formula=['0.6'],
                                  font=Font(color=ExcelStyles.RED, bold=True))

            worksheet.conditional_formatting.add(range_str, green_rule)
            worksheet.conditional_formatting.add(range_str, orange_rule)
            worksheet.conditional_formatting.add(range_str, red_rule)

    # 가치 조정 종합점수
    score_columns = ['ValuationAdjustedScore']
    for score_col in score_columns:
        if score_col in col_mapping:
            col_letter = get_column_letter(col_mapping[score_col])
            range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

            # 초록색: 0.7 (매우 매력적)
            green_rule = CellIsRule(operator='greaterThan', formula=['0.7'],
                                    font=Font(color=ExcelStyles.GREEN, bold=True))
            # 주황색: 0.6-0.7점 (보통)
            orange_rule = CellIsRule(operator='between', formula=['0.6', '0.7'],
                                     font=Font(color=ExcelStyles.ORANGE, bold=True))
            # 빨간색: 0.6점 미만 (약함)
            red_rule = CellIsRule(operator='lessThan', formula=['0.6'],
                                  font=Font(color=ExcelStyles.RED, bold=True))

            worksheet.conditional_formatting.add(range_str, green_rule)
            worksheet.conditional_formatting.add(range_str, orange_rule)
            worksheet.conditional_formatting.add(range_str, red_rule)

    # 현대적 버핏 점수
    score_columns = ['ModernBuffettScore']
    for score_col in score_columns:
        if score_col in col_mapping:
            col_letter = get_column_letter(col_mapping[score_col])
            range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

            # 초록색: 0.8점 이상 (강함)
            green_rule = CellIsRule(operator='greaterThan', formula=['0.8'],
                                    font=Font(color=ExcelStyles.GREEN, bold=True))
            # 주황색: 0.6-0.8점 (보통)
            orange_rule = CellIsRule(operator='between', formula=['0.6', '0.8'],
                                     font=Font(color=ExcelStyles.ORANGE, bold=True))
            # 빨간색: 0.6점 미만 (약함)
            red_rule = CellIsRule(operator='lessThan', formula=['0.6'],
                                  font=Font(color=ExcelStyles.RED, bold=True))

            worksheet.conditional_formatting.add(range_str, green_rule)
            worksheet.conditional_formatting.add(range_str, orange_rule)
            worksheet.conditional_formatting.add(range_str, red_rule)

    # 현대적 버핏 종합점수
    score_columns = ['TotalScore_Modern']
    for score_col in score_columns:
        if score_col in col_mapping:
            col_letter = get_column_letter(col_mapping[score_col])
            range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

            # 초록색: 75점 이상 (강함)
            green_rule = CellIsRule(operator='greaterThan', formula=['75'],
                                    font=Font(color=ExcelStyles.GREEN, bold=True))
            # 주황색: 65-75점 (보통)
            orange_rule = CellIsRule(operator='between', formula=['65', '75'],
                                     font=Font(color=ExcelStyles.ORANGE, bold=True))
            # 빨간색: 65점 미만 (약함)
            red_rule = CellIsRule(operator='lessThan', formula=['65'],
                                  font=Font(color=ExcelStyles.RED, bold=True))

            worksheet.conditional_formatting.add(range_str, green_rule)
            worksheet.conditional_formatting.add(range_str, orange_rule)
            worksheet.conditional_formatting.add(range_str, red_rule)


def apply_trading_conditional_formatting(worksheet, df, col_mapping, start_row):
    """
    트레이딩 스타일 조건부 서식 적용
    """
    # 상대거래량 (RVOL)
    if 'RVOL' in col_mapping:
        col_letter = get_column_letter(col_mapping['RVOL'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 1.2-5.0 (적정관심)
        green_rule = CellIsRule(operator='between', formula=['1.2', '5.0'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 0.8-1.2 또는 5.0-10.0 (주의)
        orange_rule1 = CellIsRule(operator='between', formula=['0.8', '1.2'],
                                  font=Font(color=ExcelStyles.ORANGE, bold=True))
        orange_rule2 = CellIsRule(operator='between', formula=['5.0', '10.0'],
                                  font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 0.8 미만 또는 10.0 초과 (위험)
        red_rule1 = CellIsRule(operator='lessThan', formula=['0.8'],
                               font=Font(color=ExcelStyles.RED, bold=True))
        red_rule2 = CellIsRule(operator='greaterThan', formula=['10.0'],
                               font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule1)
        worksheet.conditional_formatting.add(range_str, orange_rule2)
        worksheet.conditional_formatting.add(range_str, red_rule1)
        worksheet.conditional_formatting.add(range_str, red_rule2)

    # 평균변동성 (ATR_PCT)
    if 'ATR_PCT' in col_mapping:
        col_letter = get_column_letter(col_mapping['ATR_PCT'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 2-5% (적정변동성)
        green_rule = CellIsRule(operator='between', formula=['0.02', '0.05'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 1-2% 또는 5-10% (주의)
        orange_rule1 = CellIsRule(operator='between', formula=['0.01', '0.02'],
                                  font=Font(color=ExcelStyles.ORANGE, bold=True))
        orange_rule2 = CellIsRule(operator='between', formula=['0.05', '0.10'],
                                  font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 1% 미만 또는 10% 초과 (위험)
        red_rule1 = CellIsRule(operator='lessThan', formula=['0.01'],
                               font=Font(color=ExcelStyles.RED, bold=True))
        red_rule2 = CellIsRule(operator='greaterThan', formula=['0.10'],
                               font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule1)
        worksheet.conditional_formatting.add(range_str, orange_rule2)
        worksheet.conditional_formatting.add(range_str, red_rule1)
        worksheet.conditional_formatting.add(range_str, red_rule2)

    # 5일 수익률 (RET5)
    if 'RET5' in col_mapping:
        col_letter = get_column_letter(col_mapping['RET5'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 2% 이상 (강한모멘텀)
        green_rule = CellIsRule(operator='greaterThan', formula=['0.02'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 0-2% (약한모멘텀)
        orange_rule = CellIsRule(operator='between', formula=['0', '0.02'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 0% 미만 (하락모멘텀)
        red_rule = CellIsRule(operator='lessThan', formula=['0'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # 20일 수익률 (RET20)
    if 'RET20' in col_mapping:
        col_letter = get_column_letter(col_mapping['RET20'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 5% 이상 (강한상승)
        green_rule = CellIsRule(operator='greaterThan', formula=['0.05'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 0-5% (약한상승)
        orange_rule = CellIsRule(operator='between', formula=['0', '0.05'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 0% 미만 (하락추세)
        red_rule = CellIsRule(operator='lessThan', formula=['0'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # 트레이딩 점수들
    score_columns = ['MomentumScore', 'TrendScore', 'LiquidityScore']
    for score_col in score_columns:
        if score_col in col_mapping:
            col_letter = get_column_letter(col_mapping[score_col])
            range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

            # 초록색: 0.7점 이상 (강함)
            green_rule = CellIsRule(operator='greaterThan', formula=['0.7'],
                                    font=Font(color=ExcelStyles.GREEN, bold=True))
            # 주황색: 0.5-0.7점 (보통)
            orange_rule = CellIsRule(operator='between', formula=['0.5', '0.7'],
                                     font=Font(color=ExcelStyles.ORANGE, bold=True))
            # 빨간색: 0.5점 미만 (약함)
            red_rule = CellIsRule(operator='lessThan', formula=['0.5'],
                                  font=Font(color=ExcelStyles.RED, bold=True))

            worksheet.conditional_formatting.add(range_str, green_rule)
            worksheet.conditional_formatting.add(range_str, orange_rule)
            worksheet.conditional_formatting.add(range_str, red_rule)

    score_columns = ['VolatilityScore']
    for score_col in score_columns:
        if score_col in col_mapping:
            col_letter = get_column_letter(col_mapping[score_col])
            range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

            # 초록색: 0.6-0.8 (적정관심)
            green_rule = CellIsRule(operator='between', formula=['0.6', '0.8'],
                                    font=Font(color=ExcelStyles.GREEN, bold=True))
            # 주황색: 0.4-0.6 (주의)
            orange_rule = CellIsRule(operator='between', formula=['0.4', '0.6'],
                                     font=Font(color=ExcelStyles.ORANGE, bold=True))
            # 빨간색: 0.4 미만 또는 0.8 초과 (위험)
            red_rule1 = CellIsRule(operator='lessThan', formula=['0.4'],
                                   font=Font(color=ExcelStyles.RED, bold=True))
            red_rule2 = CellIsRule(operator='greaterThan', formula=['0.8'],
                                   font=Font(color=ExcelStyles.RED, bold=True))

            worksheet.conditional_formatting.add(range_str, green_rule)
            worksheet.conditional_formatting.add(range_str, orange_rule)
            worksheet.conditional_formatting.add(range_str, red_rule1)
            worksheet.conditional_formatting.add(range_str, red_rule2)

    # 토탈 점수
    score_columns = ['TotalScore']
    for score_col in score_columns:
        if score_col in col_mapping:
            col_letter = get_column_letter(col_mapping[score_col])
            range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

            # 초록색: 70점 이상 (강함)
            green_rule = CellIsRule(operator='greaterThan', formula=['70'],
                                    font=Font(color=ExcelStyles.GREEN, bold=True))
            # 주황색: 50-70점 (보통)
            orange_rule = CellIsRule(operator='between', formula=['50', '70'],
                                     font=Font(color=ExcelStyles.ORANGE, bold=True))
            # 빨간색: 50점 미만 (약함)
            red_rule = CellIsRule(operator='lessThan', formula=['50'],
                                  font=Font(color=ExcelStyles.RED, bold=True))

            worksheet.conditional_formatting.add(range_str, green_rule)
            worksheet.conditional_formatting.add(range_str, orange_rule)
            worksheet.conditional_formatting.add(range_str, red_rule)

    # RSI_14 조건부 서식
    if 'RSI_14' in col_mapping:
        col_letter = get_column_letter(col_mapping['RSI_14'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 30-70 (이상적)
        green_rule = CellIsRule(operator='between', formula=['30', '70'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 20-30 또는 70-80 (주의)
        orange_rule1 = CellIsRule(operator='between', formula=['20', '30'],
                                  font=Font(color=ExcelStyles.ORANGE, bold=True))
        orange_rule2 = CellIsRule(operator='between', formula=['70', '80'],
                                  font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 20 미만 또는 80 초과 (위험)
        red_rule1 = CellIsRule(operator='lessThan', formula=['20'],
                               font=Font(color=ExcelStyles.RED, bold=True))
        red_rule2 = CellIsRule(operator='greaterThan', formula=['80'],
                               font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule1)
        worksheet.conditional_formatting.add(range_str, orange_rule2)
        worksheet.conditional_formatting.add(range_str, red_rule1)
        worksheet.conditional_formatting.add(range_str, red_rule2)

    # MACD_Histogram 조건부 서식
    if 'MACD_Histogram' in col_mapping:
        col_letter = get_column_letter(col_mapping['MACD_Histogram'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 양수 (상승 모멘텀)
        green_rule = CellIsRule(operator='greaterThan', formula=['0'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 빨간색: 음수 (하락 모멘텀)
        red_rule = CellIsRule(operator='lessThan', formula=['0'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)

    # BB_Position 조건부 서식
    if 'BB_Position' in col_mapping:
        col_letter = get_column_letter(col_mapping['BB_Position'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 0.3-0.7 (이상적)
        green_rule = CellIsRule(operator='between', formula=['0.3', '0.7'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 0.2-0.3 또는 0.7-0.8 (주의)
        orange_rule1 = CellIsRule(operator='between', formula=['0.2', '0.3'],
                                  font=Font(color=ExcelStyles.ORANGE, bold=True))
        orange_rule2 = CellIsRule(operator='between', formula=['0.7', '0.8'],
                                  font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 0.2 미만 또는 0.8 초과 (위험)
        red_rule1 = CellIsRule(operator='lessThan', formula=['0.2'],
                               font=Font(color=ExcelStyles.RED, bold=True))
        red_rule2 = CellIsRule(operator='greaterThan', formula=['0.8'],
                               font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule1)
        worksheet.conditional_formatting.add(range_str, orange_rule2)
        worksheet.conditional_formatting.add(range_str, red_rule1)
        worksheet.conditional_formatting.add(range_str, red_rule2)

    # High_52W_Ratio 조건부 서식
    if 'High_52W_Ratio' in col_mapping:
        col_letter = get_column_letter(col_mapping['High_52W_Ratio'])
        range_str = f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}"

        # 초록색: 0.7-0.95 (이상적)
        green_rule = CellIsRule(operator='between', formula=['0.7', '0.95'],
                                font=Font(color=ExcelStyles.GREEN, bold=True))
        # 주황색: 0.5-0.7 (주의)
        orange_rule = CellIsRule(operator='between', formula=['0.5', '0.7'],
                                 font=Font(color=ExcelStyles.ORANGE, bold=True))
        # 빨간색: 0.5 미만 (위험)
        red_rule = CellIsRule(operator='lessThan', formula=['0.5'],
                              font=Font(color=ExcelStyles.RED, bold=True))

        worksheet.conditional_formatting.add(range_str, green_rule)
        worksheet.conditional_formatting.add(range_str, orange_rule)
        worksheet.conditional_formatting.add(range_str, red_rule)


def apply_excel_styling(writer, sheet_name, df, is_summary=False):
    """
    엑셀 시트에 스타일링 적용 (행 교차 색상 + 숫자 포맷팅 + 조건부 서식)
    """
    try:
        worksheet = writer.sheets[sheet_name]

        # 헤더 행 스타일 적용
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = ExcelStyles.HEADER_FILL
            cell.font = ExcelStyles.HEADER_FONT
            cell.alignment = ExcelStyles.CENTER_ALIGN
            cell.border = ExcelStyles.THIN_BORDER

        # 데이터 행에 교차 색상 적용
        for row in range(2, len(df) + 2):  # 2행부터 시작 (헤더 제외)
            is_even_row = (row % 2 == 0)

            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)

                # 행 배경색 교차 적용
                if is_even_row:
                    cell.fill = ExcelStyles.LIGHT_FILL
                else:
                    cell.fill = ExcelStyles.DARK_FILL

                # 폰트 및 정렬
                cell.font = ExcelStyles.NORMAL_FONT
                cell.border = ExcelStyles.THIN_BORDER

                # 숫자 컬럼은 오른쪽 정렬, 텍스트는 왼쪽 정렬
                col_name = df.columns[col - 1]
                if any(keyword in col_name.lower() for keyword in
                       ['score', 'pct', 'yield', 'ratio', 'price', 'value', 'roe', 'pe', 'pb', 'return', 'vol',
                        'margin', 'cap', 'sma']):
                    cell.alignment = ExcelStyles.RIGHT_ALIGN
                else:
                    cell.alignment = ExcelStyles.LEFT_ALIGN

        # 숫자 포맷팅 적용
        apply_number_formatting(worksheet, df)

        # 조건부 서식 적용
        apply_enhanced_conditional_formatting(worksheet, df, sheet_name)

        # 컬럼 너비 자동 조정
        for col_idx, col_name in enumerate(df.columns, 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)

            # 헤더 길이 확인
            max_length = max(max_length, len(str(col_name)))

            # 데이터 길이 확인 (포맷팅된 값 고려)
            for row in range(2, len(df) + 2):
                try:
                    cell_value = worksheet[f"{col_letter}{row}"].value
                    if cell_value is not None:
                        # 숫자 포맷팅을 고려한 길이 계산
                        if any(keyword in col_name.lower() for keyword in ['pct', 'yield', 'ratio']):
                            # 백분율: 숫자 * 100 + 3자리 (기호 및 소수점 고려)
                            display_length = len(f"{float(cell_value) * 100:.1f}%") if isinstance(cell_value, (int,
                                                                                                               float)) else len(
                                str(cell_value))
                        elif any(keyword in col_name.lower() for keyword in ['price', 'value']):
                            # 통화: 숫자 길이 + 3자리 (소수점 및 기호)
                            display_length = len(f"{float(cell_value):.2f}") if isinstance(cell_value,
                                                                                           (int, float)) else len(
                                str(cell_value))
                        else:
                            display_length = len(str(cell_value))
                        max_length = max(max_length, display_length)
                except:
                    pass

            adjusted_width = min(max_length + 3, 25)  # 최대 25로 제한
            worksheet.column_dimensions[col_letter].width = adjusted_width

        # 첫 번째 행과 첫 번째 열(Ticker) 고정
        # B2 셀을 기준으로 고정하면 A열(1열)과 1행이 고정됨
        worksheet.freeze_panes = 'B2'

        print(f"   ✅ {sheet_name} 시트 스타일링 적용 완료 (Ticker 열 고정)")

    except Exception as e:
        print(f"   ⚠️ {sheet_name} 시트 스타일링 중 오류: {e}")


def clean_buffett_columns(df, profile_name=None):
    """
    버핏 관련 결과에서 불필요한 컬럼 제거
    """
    # 제거할 컬럼들
    columns_to_remove = ['CreatedAtUTC', 'Source', 'Debt_to_Equity', 'BuybackYield', 'P_FFO', '_OpMarginUse']

    # 실제 존재하는 컬럼만 제거
    existing_columns_to_remove = [col for col in columns_to_remove if col in df.columns]

    if existing_columns_to_remove:
        print(f"🔧 불필요한 컬럼 제거: {existing_columns_to_remove}")
        df = df.drop(columns=existing_columns_to_remove)
    # 프로파일별 컬럼 순서 정의
    buffett_lite_strict_order = [
        'Ticker', 'Name', 'Sector', 'Industry', 'Price', 'FairValue_Composite',
        'Discount_Pct', 'DollarVol($M)', 'PE', 'PEG', 'SMA20', 'SMA50', 'ATR_PCT', 'RVOL',
        'RET5', 'RET20', 'MktCap($B)', 'RevYoY', 'OpMarginTTM',
        'OperatingMargins(info)', 'ROE(info)', 'EV_EBITDA', 'FCF_Yield', 'PB',
        'DivYield', 'PayoutRatio', 'FairValue_DCF', 'FairValue_Relative',
        'FairValue_DDM', 'FairValue_Graham', 'ROE_5Y_Avg',
        'GrowthScore', 'QualityScore', 'ValueScore', 'CatalystScore',
        'TotalScore', 'ValuationAdjustedScore'
    ]

    modern_buffett_order = [
        'Ticker', 'Name', 'Sector', 'Industry', 'Price', 'FairValue_Composite',
        'Discount_Pct', 'DollarVol($M)', 'PE', 'PEG', 'SMA20', 'SMA50', 'ATR_PCT', 'RVOL',
        'RET5', 'RET20', 'MktCap($B)', 'RevYoY', 'OpMarginTTM',
        'OperatingMargins(info)', 'ROE(info)', 'EV_EBITDA', 'FCF_Yield', 'PB',
        'DivYield', 'PayoutRatio', 'FairValue_DCF', 'FairValue_Relative',
        'FairValue_DDM', 'FairValue_Graham', 'ROE_5Y_Avg',
        'GrowthScore', 'QualityScore', 'ValueScore', 'CatalystScore',
        'TotalScore', 'ModernBuffettScore', 'TotalScore_Modern'
    ]

    # 프로파일별 순서 적용
    if profile_name == "buffett_lite" or profile_name == "buffett_strict":
        # 존재하는 컬럼만 선택
        existing_columns = [col for col in buffett_lite_strict_order if col in df.columns]
        # 기존에 있지만 순서에 없는 컬럼들은 뒤에 추가
        extra_columns = [col for col in df.columns if col not in existing_columns]
        df = df[existing_columns + extra_columns]
        print(f"🔧 {profile_name} 컬럼 순서 적용 완료")

    elif profile_name == "modern_buffett":
        # 존재하는 컬럼만 선택
        existing_columns = [col for col in modern_buffett_order if col in df.columns]
        # 기존에 있지만 순서에 없는 컬럼들은 뒤에 추가
        extra_columns = [col for col in df.columns if col not in existing_columns]
        df = df[existing_columns + extra_columns]
        print(f"🔧 {profile_name} 컬럼 순서 적용 완료")

    return df


def preprocess_data_for_display(results):
    """
    표시용 데이터 전처리 (백분율 변환 등)
    """
    processed_results = {}

    for profile_name, df in results.items():
        if df.empty:
            processed_results[profile_name] = df
            continue

        df_display = df.copy()

        # 백분율 컬럼을 0-1 범위로 변환 (엑셀에서 % 표시용)
        percent_columns = [
            'Discount_Pct', 'DivYield', 'ROE(info)', 'RevYoY',
            'OpMarginTTM', 'OperatingMargins(info)', 'FCF_Yield',
            'PayoutRatio', 'ATR_PCT', 'RET5', 'RET20'
        ]

        for col in percent_columns:
            if col in df_display.columns:
                # 현재 값이 0.15 (15%) 형태인지, 15 형태인지 확인
                sample_val = df_display[col].iloc[0] if len(df_display) > 0 else 0
                if sample_val > 1:  # 15 형태라면
                    df_display[col] = df_display[col] / 100.0
                # 0.15 형태는 그대로 유지 (엑셀에서 % 포맷으로 표시)

        # 점수 컬럼을 0-100에서 0-1로 변환 (선택사항)
        score_columns = [col for col in df_display.columns if 'Score' in col and col != 'TotalScore']
        for col in score_columns:
            if col in df_display.columns:
                sample_val = df_display[col].iloc[0] if len(df_display) > 0 else 0
                if sample_val > 1:  # 0-100 점수라면
                    df_display[col] = df_display[col] / 100.0

        processed_results[profile_name] = df_display

    return processed_results


def apply_conditional_formatting(worksheet, df, start_row=2):
    """
    조건부 서식 적용 (포맷팅된 값에 맞게 수정)
    """
    try:
        from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

        # 컬럼 인덱스 찾기
        col_mapping = {col: idx + 1 for idx, col in enumerate(df.columns)}

        # TotalScore 컬럼에 색상 스케일 적용
        if 'TotalScore' in col_mapping:
            col_letter = get_column_letter(col_mapping['TotalScore'])
            color_scale_rule = ColorScaleRule(
                start_type='num', start_value=0, start_color='FF0000',  # 빨강
                mid_type='num', mid_value=50, mid_color='FFFF00',  # 노랑
                end_type='num', end_value=100, end_color='00FF00'  # 초록
            )
            worksheet.conditional_formatting.add(
                f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}",
                color_scale_rule
            )

        # Discount_Pct 컬럼에 조건부 서식 (0-1 범위로 가정)
        if 'Discount_Pct' in col_mapping:
            col_letter = get_column_letter(col_mapping['Discount_Pct'])

            # 양수(할인)는 초록색 (0.08 = 8% 이상)
            positive_rule = CellIsRule(operator='greaterThan', formula=['0.08'],
                                       stopIfTrue=True, font=Font(color='006600', bold=True))

            # 적정 할인 (5-8%)는 파란색
            medium_rule = CellIsRule(operator='between', formula=['0.05', '0.08'],
                                     stopIfTrue=True, font=Font(color='0000FF'))

            # 음수(프리미엄)는 빨간색
            negative_rule = CellIsRule(operator='lessThan', formula=['0'],
                                       stopIfTrue=True, font=Font(color='FF0000', bold=True))

            worksheet.conditional_formatting.add(
                f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}",
                positive_rule
            )
            worksheet.conditional_formatting.add(
                f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}",
                medium_rule
            )
            worksheet.conditional_formatting.add(
                f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}",
                negative_rule
            )

        # ROE 컬럼 조건부 서식 (0-1 범위로 가정)
        if 'ROE(info)' in col_mapping:
            col_letter = get_column_letter(col_mapping['ROE(info)'])

            high_roe_rule = CellIsRule(operator='greaterThan', formula=['0.15'],
                                       stopIfTrue=True, font=Font(color='006600', bold=True))

            medium_roe_rule = CellIsRule(operator='between', formula=['0.10', '0.15'],
                                         stopIfTrue=True, font=Font(color='0000FF'))

            worksheet.conditional_formatting.add(
                f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}",
                high_roe_rule
            )
            worksheet.conditional_formatting.add(
                f"{col_letter}{start_row}:{col_letter}{len(df) + start_row - 1}",
                medium_roe_rule
            )

        print("   ✅ 조건부 서식 적용 완료")

    except Exception as e:
        print(f"   ⚠️ 조건부 서식 적용 중 오류: {e}")


class EnhancedValuationModels:
    """개선된 기관 스타일 적정가 계산 클래스 (섹터별 차별화)"""

    # 섹터별 성장률 가정 (미국 시장 기준)
    SECTOR_GROWTH_RATES = {
        "technology": 0.10,
        "healthcare": 0.08,
        "financial services": 0.06,
        "consumer defensive": 0.05,
        "energy": 0.04,
        "utilities": 0.03,
        "real estate": 0.04,
        "industrials": 0.06,
        "communication services": 0.07,
        "basic materials": 0.05,
        "consumer cyclical": 0.06
    }

    SECTOR_DISCOUNT_RATES = {
        "technology": 0.11,
        "healthcare": 0.09,
        "financial services": 0.08,
        "consumer defensive": 0.07,
        "energy": 0.09,
        "utilities": 0.06,
        "real estate": 0.08,
        "industrials": 0.08,
        "communication services": 0.09,
        "basic materials": 0.09,
        "consumer cyclical": 0.10
    }

    @staticmethod
    def sector_aware_dcf_valuation(row, terminal_rate=0.02):
        """
        섹터별 차별화된 DCF 모델
        """
        try:
            sector = str(row.get('Sector') or '').lower()
            growth_rate = EnhancedValuationModels.SECTOR_GROWTH_RATES.get(sector, 0.06)
            discount_rate = EnhancedValuationModels.SECTOR_DISCOUNT_RATES.get(sector, 0.09)

            # 현재 EPS 계산
            current_eps = row['Price'] / row['PE'] if row['PE'] and row['PE'] > 0 else 0

            if current_eps <= 0:
                return None

            # 10년간 예측
            years = 10
            future_eps = [current_eps * ((1 + growth_rate) ** i) for i in range(1, years + 1)]

            # 현금흐름 할인
            discounted_eps = [eps / ((1 + discount_rate) ** i) for i, eps in enumerate(future_eps, 1)]

            # 터미널 가치
            terminal_eps = future_eps[-1] * (1 + terminal_rate)
            terminal_value = terminal_eps / (discount_rate - terminal_rate)
            discounted_terminal = terminal_value / ((1 + discount_rate) ** years)

            return sum(discounted_eps) + discounted_terminal

        except Exception:
            return None

    @staticmethod
    def enhanced_relative_valuation(df, target_row):
        """
        개선된 동종업체 비교 (아웃라이어 강건성 향상)
        """
        try:
            sector = target_row['Sector']
            price = target_row['Price']

            # 동일 섹터 필터링
            sector_peers = df[df['Sector'] == sector]

            if len(sector_peers) < 3:  # 최소 3개 이상의 동종사 필요
                return None

            valuations = []

            # PER 비교 (아웃라이어 제거)
            if pd.notna(target_row['PE']) and target_row['PE'] > 0:
                sector_pe_clean = sector_peers['PE'][(sector_peers['PE'] > 0) & (sector_peers['PE'] < 100)]
                if len(sector_pe_clean) >= 3:
                    sector_median_pe = sector_pe_clean.median()
                    eps = price / target_row['PE']
                    pe_fair_value = sector_median_pe * eps
                    valuations.append(pe_fair_value)

            # PBR 비교
            if pd.notna(target_row['PB']) and target_row['PB'] > 0:
                sector_pb_clean = sector_peers['PB'][(sector_peers['PB'] > 0) & (sector_peers['PB'] < 20)]
                if len(sector_pb_clean) >= 3:
                    sector_median_pb = sector_pb_clean.median()
                    bps = price / target_row['PB']
                    pb_fair_value = sector_median_pb * bps
                    valuations.append(pb_fair_value)

            # EV/EBITDA 비교
            if pd.notna(target_row['EV_EBITDA']) and target_row['EV_EBITDA'] > 0:
                sector_ev_clean = sector_peers['EV_EBITDA'][
                    (sector_peers['EV_EBITDA'] > 0) & (sector_peers['EV_EBITDA'] < 30)]
                if len(sector_ev_clean) >= 3:
                    sector_median_ev_ebitda = sector_ev_clean.median()
                    ev_fair_value = price * (sector_median_ev_ebitda / target_row['EV_EBITDA'])
                    valuations.append(ev_fair_value)

            # P/FFO 비교 (리츠)
            if pd.notna(target_row.get('P_FFO')) and target_row.get('P_FFO', 0) > 0:
                sector_pffo_clean = sector_peers['P_FFO'][(sector_peers['P_FFO'] > 0) & (sector_peers['P_FFO'] < 25)]
                if len(sector_pffo_clean) >= 3:
                    sector_median_p_ffo = sector_pffo_clean.median()
                    ffo_fair_value = price * (sector_median_p_ffo / target_row['P_FFO'])
                    valuations.append(ffo_fair_value)

            if not valuations:
                return None

            # 가중평균 적용 (PER에 더 높은 가중치)
            if len(valuations) >= 2:
                weights = [0.4] + [0.6 / (len(valuations) - 1)] * (len(valuations) - 1)
                return np.average(valuations, weights=weights)
            else:
                return valuations[0]

        except Exception:
            return None

    @staticmethod
    def enhanced_dividend_discount_model(row, required_return=0.08):
        """
        개선된 배당할인모델 (배당 지속성 고려)
        """
        try:
            div_yield = row.get('DivYield', 0)
            if not div_yield or div_yield <= 0:
                return None

            # 배당성향 확인 (지속가능성)
            payout_ratio = row.get('PayoutRatio', 0)
            if payout_ratio > 0.8:  # 배당성향이 80% 초과면 위험
                return None

            current_dividend = row['Price'] * div_yield
            growth_rate = min(0.06, row.get('RevYoY', 0.03) * 0.4)  # 더 보수적 성장률

            # 고든 성장모델
            if growth_rate >= required_return:
                growth_rate = required_return - 0.01

            fair_value = current_dividend * (1 + growth_rate) / (required_return - growth_rate)
            return fair_value

        except Exception:
            return None

    @staticmethod
    def enhanced_graham_number(row):
        """
        개선된 그레이엄 넘버 (현대 시장 반영)
        """
        try:
            eps = row['Price'] / row['PE'] if row['PE'] and row['PE'] > 0 else 0
            bps = row['Price'] / row['PB'] if row['PB'] and row['PB'] > 0 else 0

            if eps <= 0 or bps <= 0:
                return None

            # 현대 시장 반영하여 계수 조정
            graham_val = math.sqrt(25 * eps * bps)  # 22.5 → 25로 조정
            return graham_val

        except Exception:
            return None


def calculate_enhanced_fair_value(df):
    """
    개선된 종합 적정가 계산 (섹터 인식 + 가중치 최적화)
    """
    fair_value_data = []

    for idx, row in df.iterrows():
        valuations = []
        weights = []

        # 다양한 모델로 적정가 계산 (섹터 인식)
        dcf_val = EnhancedValuationModels.sector_aware_dcf_valuation(row)
        if dcf_val:
            valuations.append(dcf_val)
            weights.append(0.35)  # DCF 가중치 높임

        rel_val = EnhancedValuationModels.enhanced_relative_valuation(df, row)
        if rel_val:
            valuations.append(rel_val)
            weights.append(0.40)  # 상대평가 가장 높은 가중치

        ddm_val = EnhancedValuationModels.enhanced_dividend_discount_model(row)
        if ddm_val:
            valuations.append(ddm_val)
            weights.append(0.15)  # 배당주에만 적용

        graham_val = EnhancedValuationModels.enhanced_graham_number(row)
        if graham_val:
            valuations.append(graham_val)
            weights.append(0.10)  # 기본 가치평가

        # 적정가 가중평균 계산
        if valuations:
            if len(valuations) == len(weights):
                fair_value = np.average(valuations, weights=weights)
            else:
                # 가중치가 없는 경우 동일 가중치 적용
                fair_value = np.mean(valuations)
        else:
            fair_value = None

        # 현재가 대비 할인/프리미엄률
        current_price = row['Price']
        if fair_value and current_price > 0:
            discount_pct = (fair_value - current_price) / current_price * 100
        else:
            discount_pct = None

        fair_value_data.append({
            'FairValue_DCF': dcf_val,
            'FairValue_Relative': rel_val,
            'FairValue_DDM': ddm_val,
            'FairValue_Graham': graham_val,
            'FairValue_Composite': fair_value,
            'Discount_Pct': discount_pct
        })

    return pd.DataFrame(fair_value_data, index=df.index)


# 섹터 상수 정의
FIN_SECTORS = {"banks", "financial", "insurance", "capital markets", "financial services"}
REIT_SECTORS = {"reit", "real estate", "property"}
CYCLICAL_SECTORS = {"energy", "materials", "industrials", "consumer cyclical"}
DEFENSIVE_SECTORS = {"utilities", "consumer defensive", "healthcare", "consumer staples"}

# 통합 CONFIG 설정 (현대적 버핏 철학 반영 + 엄격한 기준)
# 개선된 CONFIG 설정 (현실적인 미국 주식 기준)
CONFIG = {
    "DETAILS_CACHE_FILE": "details_cache_us_all.csv",  # 새로운 캐시 파일 사용

    # 버핏형 필터 (기술적 지표 반영)
    "MIN_MKTCAP": 500_000_000,
    "MIN_PRICE": 5.0,
    "MIN_DOLLAR_VOLUME": 5_000_000,
    "HARD_PE_MAX": 25.0,
    "MIN_REV_TTM_YOY_HF": 0.02,
    "MIN_OP_MARGIN_HF": 0.08,
    "MAX_DEBT_EQUITY": 1.5,
    "MIN_ROE_HF": 0.10,

    # 기술적 지표 필터 (버핏형에 추가)
    "BUFFETT_TECHNICAL": {
        "MAX_ATR_PCT": 0.08,  # 과도한 변동성 제한
        "MIN_RSI_14": 30,  # 과매도 상태 회피
        "MAX_RSI_14": 70,  # 과매수 상태 회피
        "MIN_BB_POSITION": 0.2,  # 지나치게 낮은 밴드 위치 제한
        "MAX_BB_POSITION": 0.8,  # 지나치게 높은 밴드 위치 제한
        "MIN_52W_RATIO": 0.6,  # 52주 저가 대비 너무 낮은 주가 제한
    },

    # 추가 필터 설정
    "OP_MARGIN_EXEMPT_SECTORS": FIN_SECTORS,
    "MIN_DISCOUNT_PCT": 8.0,  # 8% 할인 (더 현실적)
    "MAX_DISCOUNT_PCT": 50.0,  # 최대 50% 할인 (지나치게 높은 할인 제외)

    # 현대적 버핏 필터 (현실적으로 조정)
    "MODERN_BUFFETT": {
        "MIN_MKTCAP": 2_000_000_000,
        "MIN_PRICE": 10.0,
        "MIN_DOLLAR_VOLUME": 10_000_000,
        "MIN_OP_MARGIN_HF": 0.12,
        "MIN_REV_TTM_YOY_HF": 0.06,
        "MAX_DEBT_EQUITY": 1.0,
        "MIN_ROE_HF": 0.15,
        "HARD_PE_MAX": 22.0,
        "MIN_DISCOUNT_PCT": 10.0,
        "MAX_DISCOUNT_PCT": 40.0,
        "MIN_MOAT_SCORE": 0.65,
    },

    # 계층적 접근: 시가총액별 차등 조건
    "MARKET_CAP_TIERS": {
        "large_cap": {  # 100억 달러 이상
            "MIN_MKTCAP": 10_000_000_000,
            "MIN_ROE": 0.12,
            "MIN_OP_MARGIN": 0.10,
            "MAX_DEBT_EQUITY": 1.2
        },
        "mid_cap": {  # 5억~100억 달러
            "MIN_MKTCAP": 500_000_000,
            "MIN_ROE": 0.15,
            "MIN_OP_MARGIN": 0.08,
            "MAX_DEBT_EQUITY": 1.5
        },
        "small_cap": {  # 5억 달러 미만
            "MIN_MKTCAP": 100_000_000,
            "MIN_ROE": 0.18,
            "MIN_OP_MARGIN": 0.06,
            "MAX_DEBT_EQUITY": 1.8
        }
    },

    # 트레이딩 필터 (새로운 기술적 지표 반영)
    "SWING_FILTERS": {
        "MIN_PRICE": 5.0,
        "MIN_DOLLAR_VOLUME": 3_000_000,
        "MIN_RVOL": 1.1,
        "ATR_PCT_RANGE": [0.015, 0.15],
        "RSI_RANGE": [30, 70],  # RSI 필터 추가
        "MACD_CONDITION": "positive",  # MACD 양수 조건
        "BB_CONDITION": "middle",  # 볼린저밴드 중간 위치 선호
        "MIN_52W_RATIO": 0.7,  # 52주 저가 대비 70% 이상
        "MIN_RET20": -0.05  # 하락 제한 완화
    },

    "DAY_FILTERS": {
        "MIN_PRICE": 5.0,
        "MIN_DOLLAR_VOLUME": 10_000_000,
        "MIN_RVOL": 1.5,
        "ATR_PCT_RANGE": [0.025, 0.25],
        "RSI_RANGE": [40, 80],  # 데이트레이딩은 더 넓은 RSI 범위
        "MACD_CONDITION": "any",
        "BB_CONDITION": "any",
        "MIN_RET5": 0.02
    },

    # 점수 가중치 개선 (기술적 지표 반영)
    "W_GROWTH": 0.20,
    "W_QUALITY": 0.35,
    "W_VALUE": 0.35,
    "W_CATALYST": 0.10,

    # 트레이딩 점수 가중치 개선
    "TRADING_WEIGHTS": {
        "swing": {
            "momentum": 0.30,  # 모멘텀 (기존 0.45에서 감소)
            "trend": 0.25,  # 트렌드 (유지)
            "liquidity": 0.20,  # 유동성 (유지)
            "volatility": 0.10,  # 변동성 (유지)
            "technical": 0.15  # 새로운 기술적 지표 가중치 추가
        },
        "daytrade": {
            "momentum": 0.25,  # 모멘텀 (기존 0.30에서 감소)
            "trend": 0.10,  # 트렌드 (유지)
            "liquidity": 0.35,  # 유동성 (기존 0.40에서 감소)
            "volatility": 0.15,  # 변동성 (기존 0.20에서 감소)
            "technical": 0.15  # 새로운 기술적 지표 가중치 추가
        }
    },
    "OUT_PREFIX": "TECH_ENHANCED_SCREENER"
}


def calculate_technical_score(row, profile_type="buffett"):
    """
    새로운 기술적 지표들을 활용한 종합 기술 점수 계산
    """
    tech_scores = []
    weights = []

    try:
        # 1. RSI 점수 (30-70 범위가 이상적)
        rsi = row.get('RSI_14')
        if rsi and not pd.isna(rsi):
            if 30 <= rsi <= 70:
                rsi_score = 1.0 - abs(rsi - 50) / 20  # 50에 가까울수록 높은 점수
            else:
                rsi_score = 0.3  # 범위 밖이면 낮은 점수
            tech_scores.append(rsi_score)
            weights.append(0.25)

        # 2. MACD 점수
        macd_histogram = row.get('MACD_Histogram')
        if macd_histogram and not pd.isna(macd_histogram):
            # 히스토그램이 양수이면 상승 모멘텀
            macd_score = 0.5 + (macd_histogram * 10)  # 정규화 필요
            macd_score = max(0.1, min(1.0, macd_score))
            tech_scores.append(macd_score)
            weights.append(0.25)

        # 3. 볼린저밴드 위치 점수
        bb_position = row.get('BB_Position')
        if bb_position and not pd.isna(bb_position):
            # 0.3-0.7 범위가 이상적 (상단/하단 너무 치우치지 않음)
            if 0.3 <= bb_position <= 0.7:
                bb_score = 1.0 - abs(bb_position - 0.5) / 0.2
            else:
                bb_score = 0.4
            tech_scores.append(bb_score)
            weights.append(0.20)

        # 4. 52주 고가 대비 위치 점수
        high_52w_ratio = row.get('High_52W_Ratio')
        if high_52w_ratio and not pd.isna(high_52w_ratio):
            # 0.7-0.95 범위가 이상적 (너무 고점에 있지 않으면서도 강세)
            if 0.7 <= high_52w_ratio <= 0.95:
                high_52w_score = 1.0
            elif high_52w_ratio > 0.95:
                high_52w_score = 0.8 - (high_52w_ratio - 0.95) * 4  # 고점일수록 점수 감소
            else:
                high_52w_score = high_52w_ratio  # 낮을수록 점수 낮음
            tech_scores.append(max(0.1, high_52w_score))
            weights.append(0.15)

        # 5. 모멘텀 점수 (12개월)
        momentum_12m = row.get('Momentum_12M')
        if momentum_12m and not pd.isna(momentum_12m):
            # 양의 모멘텀 선호, but 너무 높은 모멘텀은 주의
            if 0 <= momentum_12m <= 0.5:
                momentum_score = 0.5 + momentum_12m
            elif momentum_12m > 0.5:
                momentum_score = 1.0 - (momentum_12m - 0.5) * 0.5
            else:
                momentum_score = 0.3  # 음의 모멘텀
            tech_scores.append(max(0.1, momentum_score))
            weights.append(0.15)

    except Exception as e:
        print(f"기술 점수 계산 중 오류: {e}")

    # 가중평균 계산
    if tech_scores and weights:
        technical_score = np.average(tech_scores, weights=weights)
    else:
        technical_score = 0.5  # 기본값

    # 프로파일별 조정
    if profile_type == "buffett":
        # 버핏형: 기술적 지표 비중 낮춤
        technical_score = 0.3 + technical_score * 0.4
    elif profile_type == "trading":
        # 트레이딩: 기술적 지표 비중 높임
        technical_score = technical_score

    return min(1.0, max(0.1, technical_score))


def get_market_cap_tier(mktcap):
    """시가총액에 따른 티어 반환"""
    if mktcap >= 10_000_000_000:  # 100억 달러 이상
        return "large_cap"
    elif mktcap >= 500_000_000:  # 5억 달러 이상
        return "mid_cap"
    else:  # 5억 달러 미만
        return "small_cap"


# 현대적 버핏 필터링 함수들
def enhanced_buffett_modern_filter(row, cfg):
    """기술적 지표를 고려한 현대적 버핏 필터"""
    modern_cfg = cfg["MODERN_BUFFETT"]
    combined_cfg = {**cfg, **modern_cfg}

    if not enhanced_pass_buffett_base(row, combined_cfg):
        return False

    # 기술적 지표 필터 적용
    rsi = row.get('RSI_14')
    if rsi and not pd.isna(rsi):
        rsi_range = modern_cfg.get("TECH_RSI_RANGE", [35, 65])
        if not (rsi_range[0] <= rsi <= rsi_range[1]):
            return False

    # 추세 강도 확인
    trend_strength = calculate_trend_strength(row)
    min_trend = modern_cfg.get("TECH_TREND_STRENGTH", 0.6)
    if trend_strength < min_trend:
        return False

    return True


def calculate_trend_strength(row):
    """추세 강도 계산 (다양한 기술적 지표 활용)"""
    strength_components = []

    # 1. 이동평균 정렬 여부
    price = row.get('Price')
    sma20 = row.get('SMA20')
    sma50 = row.get('SMA50')

    if all(x is not None and not pd.isna(x) for x in [price, sma20, sma50]):
        if price > sma20 > sma50:
            strength_components.append(1.0)
        elif price > sma20:
            strength_components.append(0.7)
        else:
            strength_components.append(0.3)

    # 2. MACD 추세
    macd = row.get('MACD')
    macd_signal = row.get('MACD_Signal')
    if all(x is not None and not pd.isna(x) for x in [macd, macd_signal]):
        if macd > macd_signal:
            strength_components.append(0.8)
        else:
            strength_components.append(0.4)

    # 3. 볼린저밴드 추세
    bb_position = row.get('BB_Position')
    if bb_position and not pd.isna(bb_position):
        if 0.4 <= bb_position <= 0.6:
            strength_components.append(0.9)  # 중간 위치 - 강한 추세
        else:
            strength_components.append(0.6)

    if strength_components:
        return np.mean(strength_components)
    else:
        return 0.5


def enhanced_pass_buffett_base(row, cfg=CONFIG, debug=False):
    """기술적 지표를 고려한 개선된 버핏 베이스 필터"""

    # 기존 기본 필터 적용
    price = row.get("Price")
    dv = (row.get("DollarVol($M)") or 0) * 1_000_000
    if pd.isna(price) or pd.isna(dv):
        if debug: print(f"  ❌ 유동성 필터 실패: price={price}, dv={dv}")
        return False

    if price < cfg.get("MIN_PRICE", 5.0) or dv < cfg.get("MIN_DOLLAR_VOLUME", 5_000_000):
        if debug: print(f"  ❌ 최소가격/거래량 필터: price={price}, dv={dv}")
        return False

    # 시가총액 필터
    mktcap = (row.get("MktCap($B)") or 0) * 1_000_000_000
    min_mktcap = cfg.get("MIN_MKTCAP", 500_000_000)
    if mktcap and mktcap < min_mktcap:
        if debug: print(f"  ❌ 시가총액 필터: mktcap={mktcap}, min={min_mktcap}")
        return False

    # 기술적 지표 필터 적용
    tech_cfg = cfg.get("BUFFETT_TECHNICAL", {})

    # ATR 변동성 필터
    atr_pct = row.get("ATR_PCT")
    max_atr = tech_cfg.get("MAX_ATR_PCT", 0.08)
    if atr_pct and atr_pct > max_atr:
        if debug: print(f"  ❌ 변동성 필터: atr_pct={atr_pct}, max={max_atr}")
        return False

    # RSI 필터
    rsi = row.get("RSI_14")
    min_rsi = tech_cfg.get("MIN_RSI_14", 30)
    max_rsi = tech_cfg.get("MAX_RSI_14", 70)
    if rsi and (rsi < min_rsi or rsi > max_rsi):
        if debug: print(f"  ❌ RSI 필터: rsi={rsi}, range=[{min_rsi}, {max_rsi}]")
        return False

    # 볼린저밴드 위치 필터
    bb_position = row.get("BB_Position")
    min_bb = tech_cfg.get("MIN_BB_POSITION", 0.2)
    max_bb = tech_cfg.get("MAX_BB_POSITION", 0.8)
    if bb_position and (bb_position < min_bb or bb_position > max_bb):
        if debug: print(f"  ❌ 볼린저밴드 필터: bb_position={bb_position}, range=[{min_bb}, {max_bb}]")
        return False

    # 52주 고가 비율 필터
    high_52w_ratio = row.get("High_52W_Ratio")
    min_52w = tech_cfg.get("MIN_52W_RATIO", 0.6)
    if high_52w_ratio and high_52w_ratio < min_52w:
        if debug: print(f"  ❌ 52주 고가 비율 필터: high_52w_ratio={high_52w_ratio}, min={min_52w}")
        return False

    # 계층적 조건 적용
    tier = get_market_cap_tier(mktcap)
    tier_cfg = cfg["MARKET_CAP_TIERS"][tier]

    # 성장성 필터 (티어별 차등)
    rev_yoy = row.get("RevYoY")
    min_rev_yoy = cfg.get("MIN_REV_TTM_YOY_HF", 0.02)
    if (rev_yoy is None) or (rev_yoy < min_rev_yoy):
        if debug: print(f"  ❌ 성장성 필터: rev_yoy={rev_yoy}, min={min_rev_yoy}")
        return False

    # 수익성 필터 (섹터별 면제 + 티어별 차등)
    sec = str(row.get("Sector") or "").lower()
    op_margin = row.get("OpMarginTTM") or row.get("OperatingMargins(info)")
    min_op_margin = tier_cfg.get("MIN_OP_MARGIN", cfg.get("MIN_OP_MARGIN_HF", 0.08))

    if sec not in cfg.get("OP_MARGIN_EXEMPT_SECTORS", FIN_SECTORS):
        if (op_margin is None) or (op_margin < min_op_margin):
            if debug: print(f"  ❌ 수익성 필터: op_margin={op_margin}, min={min_op_margin}, sector={sec}, tier={tier}")
            return False

    # 재무건전성 필터 (티어별 차등)
    debt_equity = row.get("Debt_to_Equity")
    max_debt_equity = tier_cfg.get("MAX_DEBT_EQUITY", cfg.get("MAX_DEBT_EQUITY", 1.5))
    if debt_equity and not pd.isna(debt_equity) and debt_equity > max_debt_equity:
        if debug: print(f"  ❌ 재무건전성 필터: debt_equity={debt_equity}, max={max_debt_equity}, tier={tier}")
        return False

    # 수익성 필터 (ROE, 티어별 차등)
    roe = row.get("ROE(info)") or row.get("ROE_5Y_Avg")
    min_roe = tier_cfg.get("MIN_ROE", cfg.get("MIN_ROE_HF", 0.10))
    if roe is None or pd.isna(roe) or roe < min_roe:
        if debug: print(f"  ❌ ROE 필터: roe={roe}, min={min_roe}, tier={tier}")
        return False

    # 가치 필터 (더 유연하게)
    pe = row.get("PE")
    max_pe = cfg.get("HARD_PE_MAX", 25.0)
    if (pe is not None and not pd.isna(pe) and pe > max_pe):
        if debug: print(f"  ❌ PE 필터: pe={pe}, max={max_pe}")
        return False

    # PEG 필터 (데이터 있을 때만)
    peg = row.get("PEG")
    max_peg = cfg.get("HARD_PEG_MAX", 2.0)
    if (peg is not None and not pd.isna(peg) and peg > max_peg):
        if debug: print(f"  ❌ PEG 필터: peg={peg}, max={max_peg}")
        return False

    # 적정가 할인율 필터 (현실적으로 조정)
    discount_pct = row.get('Discount_Pct')
    min_discount = cfg.get("MIN_DISCOUNT_PCT", 8.0)
    max_discount = cfg.get("MAX_DISCOUNT_PCT", 50.0)

    if discount_pct is None or pd.isna(discount_pct):
        if debug: print(f"  ❌ 할인율 데이터 없음")
        return False

    if discount_pct < min_discount:
        if debug: print(f"  ❌ 할인율 부족: discount_pct={discount_pct}, min={min_discount}")
        return False

    if discount_pct > max_discount:
        if debug: print(f"  ⚠️ 지나친 할인율: discount_pct={discount_pct}, max={max_discount} (의심 필요)")
        # 지나친 할인율은 통과시키지만 경고

    if debug: print(f"  ✅ 모든 필터 통과! (tier: {tier})")
    return True


def has_economic_moat(row, cfg):
    """경제적 해자(competitive advantage) 확인"""
    moat_score = 0
    components = []

    # 고수익성 (지속적 높은 ROE)
    roe = row.get("ROE(info)")
    if roe and roe > cfg.get("MIN_ROE_HF", 0.15):
        components.append(1.0)
    elif roe and roe > 0.12:
        components.append(0.7)
    else:
        components.append(0.3)

    # 높은 영업이익률 (가격결정력)
    op_margin = row.get("OpMarginTTM") or row.get("OperatingMargins(info)")
    if op_margin and op_margin > 0.20:
        components.append(1.0)
    elif op_margin and op_margin > 0.15:
        components.append(0.8)
    else:
        components.append(0.4)

    # 브랜드 가치 (배당 지속성으로 간접 측정)
    div_yield = row.get("DivYield")
    if div_yield and div_yield > 0.02:
        components.append(0.9)
    elif div_yield and div_yield > 0:
        components.append(0.6)
    else:
        components.append(0.3)

    moat_score = sum(components) / len(components) if components else 0
    return moat_score >= cfg.get("MIN_MOAT_SCORE", 0.7)


def has_stable_cashflow(row, cfg):
    """안정적인 현금흐름 확인"""
    # FCF Yield 기준 충족
    fcf_yield = row.get("FCF_Yield")
    if not fcf_yield or fcf_yield < cfg.get("MIN_FCFY_HF", 0.04):
        return False

    # 부채 대비 FCF 생성능력
    debt_equity = row.get("Debt_to_Equity")
    if debt_equity and debt_equity > 0:
        fcf_to_debt = fcf_yield / debt_equity
        if fcf_to_debt < 0.05:  # 부채 대비 FCF 생성능력 부족
            return False

    return True


def passes_modern_financial_health(row, cfg):
    """현대적 재무건전성 검증"""
    # 부채비율 검증
    debt_equity = row.get("Debt_to_Equity")
    if debt_equity and debt_equity > cfg.get("MAX_DEBT_EQUITY", 0.8):
        return False

    # 유동성 비율 (간접 측정 - 현재자산/현재부채 데이터가 없을 경우 기본 통과)
    current_assets = row.get("CurrentAssets")
    current_liabilities = row.get("CurrentLiabilities")
    if current_assets and current_liabilities:
        current_ratio = current_assets / current_liabilities
        if current_ratio < cfg.get("MIN_CURRENT_RATIO", 1.5):
            return False

    return True


def build_modern_buffett_scores(df: pd.DataFrame, cfg=CONFIG):
    """현대적 버핏 철학 반영 점수 계산"""
    temp = df.copy()
    modern_cfg = cfg["MODERN_BUFFETT"]

    # 기본 점수 계산 (현대적 가중치 적용)
    temp = build_scores_buffett(temp, modern_cfg)

    # 현대적 버핏 점수 요소 추가
    modern_scores = []

    for idx, row in temp.iterrows():
        modern_score_components = []

        # 1. 경제적 해자 점수
        moat_score = 0
        if has_economic_moat(row, modern_cfg):
            moat_score = 0.9
        else:
            # 해자 요소별 점수 계산
            roe_score = min(1.0, (row.get("ROE(info)") or 0) / 0.20)
            margin_score = min(1.0, (row.get("OpMarginTTM") or 0) / 0.25)
            brand_score = 1.0 if row.get("DivYield", 0) > 0.02 else 0.5
            moat_score = (roe_score + margin_score + brand_score) / 3

        modern_score_components.append(moat_score)

        # 2. 현금흐름 안정성 점수
        fcf_stability = 1.0 if has_stable_cashflow(row, modern_cfg) else 0.3
        modern_score_components.append(fcf_stability)

        # 3. 재무건전성 점수
        health_score = 1.0 if passes_modern_financial_health(row, modern_cfg) else 0.4
        modern_score_components.append(health_score)

        # 4. 경기방어성 점수 (섹터 기반)
        sector = str(row.get("Sector") or "").lower()
        defensive_score = 0.7  # 기본값
        if any(x in sector for x in ["consumer defensive", "utilities", "healthcare"]):
            defensive_score = 0.9
        elif any(x in sector for x in ["technology", "financial"]):
            defensive_score = 0.8
        elif any(x in sector for x in ["energy", "cyclical"]):
            defensive_score = 0.5

        modern_score_components.append(defensive_score)

        modern_score = sum(modern_score_components) / len(modern_score_components)
        modern_scores.append(modern_score)

    temp["ModernBuffettScore"] = pd.Series(modern_scores, index=temp.index)

    # 종합 점수에 현대적 요소 반영
    temp["TotalScore_Modern"] = (
            temp["TotalScore"] * 0.7 +
            temp["ModernBuffettScore"] * 100 * 0.3
    )

    return temp


def create_detailed_explanation_sheets(writer):
    """
    상세한 설명 시트들 생성
    """
    buffett_explanations = [
        {
            '열 이름': 'Ticker',
            '의미': '종목코드',
            '설명': '주식 시장에서 사용하는 고유 기호',
            '적정 범위/기준': '-'
        },
        {
            '열 이름': 'Name',
            '의미': '회사명',
            '설명': '상장회사 공식 명칭',
            '적정 범위/기준': '-'
        },
        {
            '열 이름': 'Sector',
            '의미': '업종/섹터',
            '설명': '기술, 헬스케어, 금융 등 산업 분류',
            '적정 범위/기준': '-'
        },
        {
            '열 이름': 'Industry',
            '의미': '산업',
            '설명': '더 세부적인 산업 분류',
            '적정 범위/기준': '-'
        },
        {
            '열 이름': 'Price',
            '의미': '현재 주가',
            '설명': '현재 시장에서 거래되는 주식 가격',
            '적정 범위/기준': '✅ 10달러 이상 (저가주 리스크 회피)'
        },
        {
            '열 이름': 'FairValue_Composite',
            '의미': '종합 적정가',
            '설명': '4가지 가치평가 모델(DCF, 상대평가, 배당모델, 그레이엄)의 가중평균\n• DCF(35%), 상대평가(40%), 배당모델(15%), 그레이엄(10%)',
            '적정 범위/기준': '현재가보다 높을수록 좋음\n✅ 20% 이상 할인: 매우 매력적\n⚠️ 0-20% 할인: 보통\n❌ 프리미엄: 고평가'
        },
        {
            '열 이름': 'FairValue_DCF',
            '의미': 'DCF 적정가',
            '설명': '할인현금흐름 모델: 10년간 예측 현금흐름을 현재가치로 할인\n• 성장률: 섹터별 차등(기술 10%, 유틸리티 3% 등)\n• 할인율: 섹터별 차등(기술 11%, 유틸리티 6% 등)\n• 터미널가치: 2% 영구성장률 가정',
            '적정 범위/기준': '현재가보다 높을수록 좋음\n기업의 장기 내재가치 반영'
        },
        {
            '열 이름': 'FairValue_Relative',
            '의미': '상대평가 적정가',
            '설명': '동종업체 비교를 통한 적정가 (PER, PBR, EV/EBITDA, P/FFO)\n• 동일 섹터 3개 이상 기업과 비교\n• 이상치 제거(IQR 방식) 후 중간값 사용\n• PER(40%), PBR(30%), EV/EBITDA(30%) 가중평균',
            '적정 범위/기준': '시장 상대적 평가\n동종업체 대비 저평가/고평가 판단'
        },
        {
            '열 이름': 'FairValue_DDM',
            '의미': '배당할인모델 적정가',
            '설명': '배당할인모델: 미래 배당금을 현재가치로 할인\n• 현재 배당금 × (1 + 성장률) ÷ (필요수익률 - 성장률)\n• 성장률: 매출성장률의 40% 적용 (보수적)\n• 필요수익률: 8% 고정\n• 배당성향 80% 초과시 계산 제외',
            '적정 범위/기준': '배당주에만 의미있음\n✅ 안정적 배당기업 평가용'
        },
        {
            '열 이름': 'FairValue_Graham',
            '의미': '그레이엄 적정가',
            '설명': '벤저민 그레이엄의 가치공식: √(22.5 × EPS × BPS)\n• EPS: 주당순이익 (Price ÷ PE)\n• BPS: 주당순자산 (Price ÷ PB)\n• 현대 시장 반영하여 계수 22.5 → 25로 조정',
            '적정 범위/기준': '보수적인 가치평가\n✅ 저PER, 저PBR 기업에 효과적'
        },
        {
            '열 이름': 'Discount_Pct',
            '의미': '할인율',
            '설명': '종합적정가 대비 현재 주가 할인율\n• (FairValue_Composite - Price) ÷ Price × 100\n• 양수: 저평가, 음수: 고평가',
            '적정 범위/기준': '✅ 8-40%: 좋음 (안전마진 확보)\n⚠️ 0-8%: 보통\n❌ 0% 이하: 고평가 (매수 부적합)'
        },
        # ... [나머지 35개 지표 설명은 동일하게 유지] ...
        {
            '열 이름': 'MktCap($B)',
            '의미': '시가총액',
            '설명': '회사의 전체 시장 가치 (주가 × 발행주식수)',
            '적정 범위/기준': '✅ 5억$ 이상: 중형주\n✅ 20억$ 이상: 대형주'
        },
        {
            '열 이름': 'PE',
            '의미': '주가수익비율',
            '설명': '주가를 주당순이익으로 나눈 값, 낮을수록 저평가',
            '적정 범위/기준': '✅ 8-20배: 저PER\n⚠️ 20-25배: 보통\n❌ 25배 이상: 고PER'
        },
        {
            '열 이름': 'PEG',
            '의미': 'PER 성장률 배수',
            '설명': 'PER을 연간 성장률로 나눈 값, 1 이하가 이상적',
            '적정 범위/기준': '✅ 0.5-1.0: 매우 좋음\n⚠️ 1.0-1.5: 보통\n❌ 1.5 이상: 고평가'
        },
        {
            '열 이름': 'PB',
            '의미': '주가순자산비율',
            '설명': '주가를 주당순자산으로 나눈 값',
            '적정 범위/기준': '✅ 0.8-2.0: 적정\n⚠️ 2.0-3.0: 보통\n❌ 3.0 이상: 고평가'
        },
        {
            '열 이름': 'EV_EBITDA',
            '의미': '기업가치/EBITDA 비율',
            '설명': '기업 인수 비용 대비 영업이익 비율',
            '적정 범위/기준': '✅ 5-12배: 좋음\n⚠️ 12-18배: 보통\n❌ 18배 이상: 고평가'
        },
        {
            '열 이름': 'FCF_Yield',
            '의미': '자유현금흐름 수익률',
            '설명': '주가 대비 자유현금흐름 비율',
            '적정 범위/기준': '✅ 5% 이상: 우량\n⚠️ 2-5%: 보통\n❌ 2% 미만: 약함'
        },
        {
            '열 이름': 'P_FFO',
            '의미': '주당운영현금흐름 배수',
            '설명': '리츠(REITs) 평가 지표, 주가를 주당운영현금흐름으로 나눈 값',
            '적정 범위/기준': '✅ 8-15배: 적정\n⚠️ 15-20배: 보통\n❌ 20배 이상: 고평가'
        },
        {
            '열 이름': 'DivYield',
            '의미': '배당수익률',
            '설명': '주가 대비 배당금 비율',
            '적정 범위/기준': '✅ 2-6%: 적정\n⚠️ 6% 이상: 주의필요\n❌ 0%: 배당없음'
        },
        {
            '열 이름': 'PayoutRatio',
            '의미': '배당성향',
            '설명': '순이익 대비 배당금 비율',
            '적정 범위/기준': '✅ 30-60%: 적정\n⚠️ 60-80%: 주의\n❌ 80% 이상: 위험'
        },
        {
            '열 이름': 'BuybackYield',
            '의미': '자사주매수 수익률',
            '설명': '시가총액 대비 자사주매수 규모',
            '적정 범위/기준': '✅ 1-5%: 긍정적\n⚠️ 5% 이상: 과도할 수 있음'
        },
        {
            '열 이름': 'RevYoY',
            '의미': '매출 성장률',
            '설명': '전년 동기 대비 매출 증가율',
            '적정 범위/기준': '✅ 5% 이상: 강한성장\n⚠️ 2-5%: 보통성장\n❌ 2% 미만: 낮은성장'
        },
        {
            '열 이름': 'EPSYoY',
            '의미': 'EPS 성장률',
            '설명': '전년 동기 대비 주당순이익 증가율',
            '적정 범위/기준': '✅ 8% 이상: 강한성장\n⚠️ 3-8%: 보통성장\n❌ 3% 미만: 낮은성장'
        },
        {
            '열 이름': 'OpMarginTTM',
            '의미': '영업이익률',
            '설명': '매출 대비 영업이익 비율, 사업 효율성 지표',
            '적정 범위/기준': '✅ 12% 이상: 고효율\n⚠️ 8-12%: 보통\n❌ 8% 미만: 저효율'
        },
        {
            '열 이름': 'NetMarginTTM',
            '의미': '순이익률',
            '설명': '매출 대비 순이익 비율, 최종 수익성 지표',
            '적정 범위/기준': '✅ 10% 이상: 고수익\n⚠️ 5-10%: 보통\n❌ 5% 미만: 저수익'
        },
        {
            '열 이름': 'ROE(info)',
            '의미': '자기자본이익률',
            '설명': '자본 대비 순이익률, 수익성 지표',
            '적정 범위/기준': '✅ 15% 이상: 우량\n⚠️ 10-15%: 보통\n❌ 10% 미만: 약함'
        },
        {
            '열 이름': 'ROA(info)',
            '의미': '총자산이익률',
            '설명': '총자산 대비 순이익률, 자산 효율성 지표',
            '적정 범위/기준': '✅ 8% 이상: 우량\n⚠️ 5-8%: 보통\n❌ 5% 미만: 약함'
        },
        {
            '열 이름': 'ROIC(info)',
            '의미': '투하자본이익률',
            '설명': '투하자본 대비 영업이익률, 투자 효율성 지표',
            '적정 범위/기준': '✅ 12% 이상: 우량\n⚠️ 8-12%: 보통\n❌ 8% 미만: 약함'
        },
        {
            '열 이름': 'ROE_5Y_Avg',
            '의미': '5년 평균 ROE',
            '설명': '5년간 평균 자기자본이익률, 수익성의 지속성과 안정성 측정\n• 단년도 ROE보다 장기적인 수익성 파악에 유용\n• 변동성이 적고 일관된 수익성 나타냄',
            '적정 범위/기준': '✅ 20% 이상: 탁월한 수익성 지속성\n✅ 15-20%: 우량한 수익성 지속성\n⚠️ 10-15%: 보통 수준의 수익성 지속성\n❌ 10% 미만: 낮은 수익성 지속성'
        },
        {
            '열 이름': 'Debt_to_Equity',
            '의미': '부채비율',
            '설명': '자본 대비 부채 비율, 낮을수록 재무건전성 좋음',
            '적정 범위/기준': '✅ 0.5 이하: 매우 건전\n⚠️ 0.5-1.5: 보통\n❌ 1.5 이상: 위험'
        },
        {
            '열 이름': 'CurrentRatio',
            '의미': '유동비율',
            '설명': '유동자산 대비 유동부채 비율, 단기 유동성 지표',
            '적정 범위/기준': '✅ 1.5-3.0: 적정\n⚠️ 1.0-1.5: 주의\n❌ 1.0 미만: 위험'
        },
        {
            '열 이름': 'QuickRatio',
            '의미': '당좌비율',
            '설명': '당좌자산 대비 유동부채 비율, 즉시 유동성 지표',
            '적정 범위/기준': '✅ 1.0 이상: 안전\n⚠️ 0.5-1.0: 주의\n❌ 0.5 미만: 위험'
        },
        {
            '열 이름': 'InterestCoverage',
            '의미': '이자보상배수',
            '설명': '영업이익 대비 이자비용 비율, 이자지급 능력',
            '적정 범위/기준': '✅ 5배 이상: 안전\n⚠️ 3-5배: 보통\n❌ 3배 미만: 위험'
        },
        {
            '열 이름': 'OperatingMargins(info)',
            '의미': '영업이익률 (info)',
            '설명': 'yfinance 제공 영업이익률 데이터',
            '적정 범위/기준': '✅ 12% 이상: 우량\n⚠️ 8-12%: 보통'
        },
        {
            '열 이름': 'ProfitMargins(info)',
            '의미': '순이익률 (info)',
            '설명': 'yfinance 제공 순이익률 데이터',
            '적정 범위/기준': '✅ 10% 이상: 우량\n⚠️ 5-10%: 보통'
        },
        {
            '열 이름': 'GrossMargins(info)',
            '의미': '매출총이익률',
            '설명': '매출 대비 매출총이익 비율, 원가 관리 효율성',
            '적정 범위/기준': '✅ 40% 이상: 고효율\n⚠️ 20-40%: 보통\n❌ 20% 미만: 저효율'
        },
        {
            '열 이름': 'DollarVol($M)',
            '의미': '달러 거래량',
            '설명': '하루 평균 거래 대금 (백만 달러)',
            '적정 범위/기준': '✅ 10M$ 이상: 높은유동성\n⚠️ 5-10M$: 보통유동성'
        },
        {
            '열 이름': 'RVOL',
            '의미': '상대 거래량',
            '설명': '평균 대비 거래량 비율 (1.0 = 평균)',
            '적정 범위/기준': '✅ 0.8-2.0: 적정\n⚠️ 2.0 이상: 과열 가능성'
        },
        {
            '열 이름': 'ATR_PCT',
            '의미': '평균 실제 범위',
            '설명': '일일 평균 주가 변동폭 (%)',
            '적정 범위/기준': '✅ 1-5%: 안정적\n⚠️ 5-10%: 변동성 있음\n❌ 10% 이상: 고변동성'
        },
        {
            '열 이름': 'SMA20',
            '의미': '20일 이동평균',
            '설명': '단기 추세선, 20일간 평균 주가',
            '적정 범위/기준': '✅ 주가 > SMA20: 상승추세\n❌ 주가 < SMA20: 하락추세'
        },
        {
            '열 이름': 'SMA50',
            '의미': '50일 이동평균',
            '설명': '중기 추세선, 50일간 평균 주가',
            '적정 범위/기준': '✅ SMA20 > SMA50: 강한상승\n❌ SMA20 < SMA50: 약세'
        },
        {
            '열 이름': 'RET5',
            '의미': '5일 수익률',
            '설명': '최근 5일간 주가 등락율',
            '적정 범위/기준': '✅ -5%~+5%: 안정적\n⚠️ ±5-10%: 변동성 있음'
        },
        {
            '열 이름': 'RET20',
            '의미': '20일 수익률',
            '설명': '최근 20일간 주가 등락율',
            '적정 범위/기준': '✅ -10%~+15%: 안정적\n⚠️ ±15% 이상: 변동성 큼'
        },
        {
            '열 이름': 'GrowthScore',
            '의미': '성장성 점수',
            '설명': '매출 성장, 수익 성장 등 성장성 종합 점수',
            '적정 범위/기준': '✅ 70점 이상: 강한성장\n⚠️ 60-70점: 보통성장\n❌ 60점 미만: 낮은성장'
        },
        {
            '열 이름': 'QualityScore',
            '의미': '질 점수',
            '설명': '수익성, 재무건전성, 경영 효율성 종합 점수',
            '적정 범위/기준': '✅ 70점 이상: 우량기업\n⚠️ 60-70점: 보통기업\n❌ 60점 미만: 취약기업'
        },
        {
            '열 이름': 'ValueScore',
            '의미': '가치 점수',
            '설명': '저평가 정도, 다양한 가치 지표 종합 점수',
            '적정 범위/기준': '✅ 70점 이상: 저평가\n⚠️ 60-70점: 공정가치\n❌ 60점 미만: 고평가'
        },
        {
            '열 이름': 'CatalystScore',
            '의미': '촉매 점수',
            '설명': '배당, 자사주매수 등 주가 상승 촉매 요소 점수',
            '적정 범위/기준': '✅ 70점 이상: 강한촉매\n⚠️ 60-70점: 보통촉매\n❌ 60점 미만: 약한촉매'
        },
        {
            '열 이름': 'TotalScore',
            '의미': '종합 총점',
            '설명': '성장성 + 질 + 가치 + 촉매 점수의 가중합',
            '적정 범위/기준': '✅ 70점 이상: 최우량\n⚠️ 60-70점: 우량\n❌ 60점 미만: 일반'
        },
        {
            '열 이름': 'ValuationAdjustedScore',
            '의미': '가치 조정 종합점수',
            '설명': '종합 총점에 할인율을 추가 반영한 최종 점수\n• TotalScore × (1 + Discount_Pct/100)',
            '적정 범위/기준': '✅ 80점 이상: 매우매력적\n⚠️ 70-80점: 매력적\n❌ 70점 미만: 보통'
        },
        {
            '열 이름': 'ModernBuffettScore',
            '의미': '현대적 버핏 점수',
            '설명': '경제적 해자, 현금흐름 안정성 등 현대적 버핏 요소 점수\n• ROE, 영업이익률, 배당 지속성 종합',
            '적정 범위/기준': '✅ 0.8 이상: 강한해자\n⚠️ 0.6-0.8: 보통해자\n❌ 0.6 미만: 약한해자'
        },
        {
            '열 이름': 'TotalScore_Modern',
            '의미': '현대적 버핏 종합점수',
            '설명': '현대적 버핏 철학을 반영한 최종 점수\n• TotalScore(70%) + ModernBuffettScore(30%)',
            '적정 범위/기준': '✅ 75점 이상: 현대적우량\n⚠️ 65-75점: 현대적보통\n❌ 65점 미만: 일반'
        }
    ]

    buffett_df = pd.DataFrame(buffett_explanations)
    buffett_df.to_excel(writer, sheet_name='버핏_지표_설명', index=False)
    apply_excel_styling(writer, '버핏_지표_설명', buffett_df)

    print(f"   ✅ 버핏 지표 설명: {len(buffett_explanations)}개 지표 추가 (FairValue 구체화)")

    # 2. 트레이딩 프로파일 설명 시트
    swing_daytrading_explanations = [
        {
            '열 이름': 'Ticker',
            '의미': '종목코드',
            '설명': '주식 시장에서 사용하는 고유 기호',
            '적정 범위/기준': '-'
        },
        {
            '열 이름': 'Name',
            '의미': '회사명',
            '설명': '상장회사 공식 명칭',
            '적정 범위/기준': '-'
        },
        {
            '열 이름': 'Sector',
            '의미': '업종/섹터',
            '설명': '기술, 헬스케어, 금융 등 산업 분류',
            '적정 범위/기준': '고변동성 섹터(기술, 바이오) 선호'
        },
        {
            '열 이름': 'Price',
            '의미': '현재 주가',
            '설명': '현재 시장에서 거래되는 주식 가격',
            '적정 범위/기준': '✅ 5-50달러: 적정거래구간\n❌ 5달러 미만: 위험\n❌ 50달러 이상: 고가주'
        },
        {
            '열 이름': 'DollarVol($M)',
            '의미': '달러 거래량',
            '설명': '하루 평균 거래 대금 (백만 달러)',
            '적정 범위/기준': '✅ 5M$ 이상: 높은유동성\n⚠️ 1-5M$: 보통유동성\n❌ 1M$ 미만: 낮은유동성'
        },
        {
            '열 이름': 'RVOL',
            '의미': '상대 거래량',
            '설명': '평균 대비 거래량 비율 (1.0 = 평균)',
            '적정 범위/기준': '✅ 1.2-5.0: 적정관심\n⚠️ 0.8-1.2: 평균\n❌ 0.8 미만: 관심낮음'
        },
        {
            '열 이름': 'ATR_PCT',
            '의미': '평균 실제 범위',
            '설명': '일일 평균 주가 변동폭 (%)',
            '적정 범위/기준': '✅ 2-10%: 적정변동성\n⚠️ 10-15%: 고변동성\n❌ 15% 이상: 매우높은변동성'
        },
        {
            '열 이름': 'SMA20',
            '의미': '20일 이동평균',
            '설명': '단기 추세선, 20일간 평균 주가',
            '적정 범위/기준': '✅ 주가 > SMA20: 상승추세\n⚠️ 주가 ≈ SMA20: 횡보\n❌ 주가 < SMA20: 하락추세'
        },
        {
            '열 이름': 'SMA50',
            '의미': '50일 이동평균',
            '설명': '중기 추세선, 50일간 평균 주가',
            '적정 범위/기준': '✅ SMA20 > SMA50: 강한상승\n⚠️ SMA20 ≈ SMA50: 중립\n❌ SMA20 < SMA50: 약세'
        },
        {
            '열 이름': 'RET5',
            '의미': '5일 수익률',
            '설명': '최근 5일간 주가 등락율',
            '적정 범위/기준': '✅ 2-10%: 강한모멘텀\n⚠️ 0-2%: 약한모멘텀\n❌ 0% 미만: 하락모멘텀'
        },
        {
            '열 이름': 'RET20',
            '의미': '20일 수익률',
            '설명': '최근 20일간 주가 등락율',
            '적정 범위/기준': '✅ 5-25%: 강한상승\n⚠️ 0-5%: 약한상승\n❌ 0% 미만: 하락추세'
        },
        {
            '열 이름': 'MomentumScore',
            '의미': '모멘텀 점수',
            '설명': '단기 주가 추세 강도 (최근 상승력)',
            '적정 범위/기준': '✅ 0.7점 이상: 강한모멘텀\n⚠️ 0.5-0.7점: 보통모멘텀\n❌ 0.5점 미만: 약한모멘텀'
        },
        {
            '열 이름': 'TrendScore',
            '의미': '트렌드 점수',
            '설명': '장기 추세 방향성 (상승/하락/횡보)',
            '적정 범위/기준': '✅ 0.7점 이상: 강한상승추세\n⚠️ 0.5-0.7점: 약한상승/횡보\n❌ 0.5점 미만: 하락추세'
        },
        {
            '열 이름': 'LiquidityScore',
            '의미': '유동성 점수',
            '설명': '매매 용이성 (거래량, 거래대금 종합)',
            '적정 범위/기준': '✅ 0.7점 이상: 높은유동성\n⚠️ 0.5-0.7점: 보통유동성\n❌ 0.5점 미만: 낮은유동성'
        },
        {
            '열 이름': 'VolatilityScore',
            '의미': '변동성 점수',
            '설명': '적정 변동성 (너무 낮거나 높지 않은 적정 수준)',
            '적정 범위/기준': '✅ 0.6-0.8점: 이상적변동성\n⚠️ 0.4-0.6점: 높은변동성\n❌ 0.4점 미만: 매우높은변동성'
        },
        {
            '열 이름': 'TotalScore',
            '의미': '종합 총점',
            '설명': '모멘텀 + 트렌드 + 유동성 + 변동성 점수의 가중합',
            '적정 범위/기준': '✅ 70점 이상: 최우량\n⚠️ 60-70점: 우량\n❌ 60점 미만: 일반'
        }
    ]

    trading_df = pd.DataFrame(swing_daytrading_explanations)
    trading_df.to_excel(writer, sheet_name='스윙&daytrade_지표_설명', index=False)
    apply_excel_styling(writer, '스윙&daytrade_지표_설명', trading_df)

    # 3. 프로파일 비교 설명 시트
    profile_comparison = [
        {
            '프로파일': 'buffett_lite',
            '투자 스타일': '기본 가치투자',
            '보유 기간': '장기 (1-5년)',
            '목표 수익률': '연 10-15%',
            '주요 초점': '내재가치 대비 저평가 + 재무건전성',
            '위험 수준': '낮음',
            '추천 투자자': '가치투자 입문자, 안정성 추구자',
            '핵심 지표': 'Discount_Pct, ROE, PE, Debt_to_Equity'
        },
        {
            '프로파일': 'buffett_strict',
            '투자 스타일': '엄격한 가치투자',
            '보유 기간': '장기 (3-7년)',
            '목표 수익률': '연 15-20%',
            '주요 초점': '고품질 기업 + 확실한 안전마진',
            '위험 수준': '낮음-중간',
            '추천 투자자': '경험이 많은 가치투자자',
            '핵심 지표': 'Discount_Pct, ROE, OpMargin, ModernBuffettScore'
        },
        {
            '프로파일': 'modern_buffett',
            '투자 스타일': '현대적 가치투자',
            '보유 기간': '장기 (5년+)',
            '목표 수익률': '연 12-18%',
            '주요 초점': '경제적 해자 + 지속성장 가능성',
            '위험 수준': '낮음-중간',
            '추천 투자자': '워렌 버핏 철학 따르는 투자자',
            '핵심 지표': 'ModernBuffettScore, TotalScore_Modern, Sector'
        },
        {
            '프로파일': 'swing',
            '투자 스타일': '스윙트레이딩',
            '보유 기간': '중기 (수일-수주)',
            '목표 수익률': '월 5-15%',
            '주요 초점': '추세 + 모멘텀 + 기술적 분석',
            '위험 수준': '중간-높음',
            '추천 투자자': '활동적 트레이더, 기술적 분석가',
            '핵심 지표': 'MomentumScore, TrendScore, ATR_PCT, RVOL'
        },
        {
            '프로파일': 'daytrade',
            '투자 스타일': '데이트레이딩',
            '보유 기간': '단기 (당일)',
            '목표 수익률': '일 0.5-3%',
            '주요 초점': '유동성 + 변동성 + 단기 모멘텀',
            '위험 수준': '높음',
            '추천 투자자': '전문 트레이더, 단기 스캘퍼',
            '핵심 지표': 'LiquidityScore, VolatilityScore, RET5, DollarVol'
        }
    ]

    profile_df = pd.DataFrame(profile_comparison)
    profile_df.to_excel(writer, sheet_name='프로파일_비교', index=False)
    apply_excel_styling(writer, '프로파일_비교', profile_df)

    # 색상 기준 설명 시트 추가
    color_guidelines = [
        {
            '색상': '🟢 초록색',
            '의미': '좋은 수치',
            '설명': '투자/트레이딩에 매우 유리한 조건',
            '예시': '할인율 8% 이상, ROE 15% 이상, PER 15배 이하 등'
        },
        {
            '색상': '🟡 주황색',
            '의미': '주의',
            '설명': '보통 수준이거나 주의가 필요한 조건',
            '예시': '할인율 0-8%, ROE 10-15%, PER 15-25배 등'
        },
        {
            '색상': '🔴 빨간색',
            '의미': '위험한 수치',
            '설명': '투자/트레이딩에 불리하거나 위험한 조건',
            '예시': '프리미엄 거래, ROE 10% 미만, PER 25배 초과 등'
        }
    ]

    color_df = pd.DataFrame(color_guidelines)
    color_df.to_excel(writer, sheet_name='색상_기준_설명', index=False)
    apply_excel_styling(writer, '색상_기준_설명', color_df)

    print("   ✅ 색상 기준 설명 시트 추가")


def create_styled_excel_output(results, filename):
    """
    상세한 설명 시트가 포함된 엑셀 파일 생성
    """
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:

        # 1. 각 프로파일 시트 저장 및 스타일링
        for profile_name, result_df in results.items():
            if not result_df.empty:
                # 이미 clean_buffett_columns에서 정리되었으므로 바로 저장
                result_df.to_excel(writer, sheet_name=profile_name[:31], index=False)

                # 기본 스타일링 적용
                apply_excel_styling(writer, profile_name[:31], result_df)

        # 2. 요약 시트 생성
        summary_data = []
        for profile_name, result_df in results.items():
            if not result_df.empty:
                if profile_name.startswith('buffett') or profile_name == 'modern_buffett':
                    stats = {
                        'Profile': profile_name,
                        'Stocks_Count': len(result_df),
                        'Avg_Discount_Pct': result_df['Discount_Pct'].mean(),
                        'Median_PE': result_df['PE'].median(),
                        'Avg_ROE': result_df['ROE(info)'].mean(),
                        'Top_Tickers': ', '.join(result_df.head(3)['Ticker'].tolist())
                    }
                else:
                    stats = {
                        'Profile': profile_name,
                        'Stocks_Count': len(result_df),
                        'Avg_RVOL': result_df['RVOL'].mean(),
                        'Avg_ATR_PCT': result_df['ATR_PCT'].mean(),
                        'Top_Tickers': ', '.join(result_df.head(3)['Ticker'].tolist())
                    }
                summary_data.append(stats)

        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            apply_excel_styling(writer, 'Summary', summary_df, is_summary=True)

        # 3. 상세한 설명 시트들 생성
        print("📚 상세한 설명 시트 생성 중...")
        create_detailed_explanation_sheets(writer)


def enhanced_valuation_screener_with_formatting():
    """
    상세한 설명이 포함된 개선된 통합 스크리너 (컬럼 순서 적용)
    """
    # 데이터 로드 및 처리
    df = load_cache(CONFIG["DETAILS_CACHE_FILE"])
    # 전체 데이터에 대해 Winsorization 적용
    df = apply_winsorization_to_whole_dataset(df)
    print("Calculating enhanced fair values with sector awareness...")
    fair_values_df = calculate_enhanced_fair_value(df)
    df = pd.concat([df, fair_values_df], axis=1)
    results = {}

    # ROE_5Y_Avg 데이터 품질 확인
    if 'ROE_5Y_Avg' in df.columns:
        roe_5y_non_null = df['ROE_5Y_Avg'].notna().sum()
        roe_info_non_null = df['ROE(info)'].notna().sum() if 'ROE(info)' in df.columns else 0
        print(f"📊 ROE 데이터 품질: ROE_5Y_Avg {roe_5y_non_null}/{len(df)}, ROE(info) {roe_info_non_null}/{len(df)}")

        # ROE_5Y_Avg가 없는 경우 ROE(info)로 대체
        if roe_5y_non_null == 0 and roe_info_non_null > 0:
            print("   🔄 ROE_5Y_Avg 데이터 없음, ROE(info)로 대체")
            df['ROE_5Y_Avg'] = df['ROE(info)']

    results = {}

    # 1. 버핏-Lite (ROE_5Y_Avg 포함)
    mask_lite = df.apply(lambda r: enhanced_pass_buffett_base(r, CONFIG), axis=1)
    raw_lite = df[mask_lite].copy()
    if not raw_lite.empty:
        scored_lite = build_scores_buffett(raw_lite, CONFIG)
        scored_lite['ValuationAdjustedScore'] = scored_lite['TotalScore'] * (
                1 + scored_lite['Discount_Pct'].fillna(0) / 100
        )
        scored_lite = scored_lite[scored_lite['TotalScore'] >= 60]

        # ROE_5Y_Avg 데이터 정리
        if 'ROE_5Y_Avg' not in scored_lite.columns and 'ROE(info)' in scored_lite.columns:
            scored_lite['ROE_5Y_Avg'] = scored_lite['ROE(info)']

        # 버핏 결과 컬럼 정리 및 순서 적용
        scored_lite = clean_buffett_columns(scored_lite, "buffett_lite")
        results["buffett_lite"] = scored_lite.sort_values("ValuationAdjustedScore", ascending=False)

    # 2. 버핏-Strict (ROE_5Y_Avg 포함)
    strict_cfg = CONFIG.copy()
    strict_cfg.update({
        "MIN_MKTCAP": 2_000_000_000,
        "MIN_PRICE": 10.0,
        "MIN_DOLLAR_VOLUME": 10_000_000,
        "MIN_DISCOUNT_PCT": 12.0,
        "MIN_OP_MARGIN_HF": 0.12,
        "MIN_REV_TTM_YOY_HF": 0.06,
        "HARD_PE_MAX": 20.0,
        "MIN_ROE_HF": 0.15,
        "MAX_DEBT_EQUITY": 1.0,
    })
    mask_strict = df.apply(lambda r: enhanced_pass_buffett_base(r, strict_cfg), axis=1)
    raw_strict = df[mask_strict].copy()
    if not raw_strict.empty:
        scored_strict = build_scores_buffett(raw_strict, strict_cfg)
        scored_strict['ValuationAdjustedScore'] = scored_strict['TotalScore'] * (
                1 + scored_strict['Discount_Pct'].fillna(0) / 100
        )
        scored_strict = scored_strict[scored_strict['TotalScore'] >= 70]

        # ROE_5Y_Avg 데이터 정리
        if 'ROE_5Y_Avg' not in scored_strict.columns and 'ROE(info)' in scored_strict.columns:
            scored_strict['ROE_5Y_Avg'] = scored_strict['ROE(info)']

        # 버핏 결과 컬럼 정리 및 순서 적용
        scored_strict = clean_buffett_columns(scored_strict, "buffett_strict")
        results["buffett_strict"] = scored_strict.sort_values("ValuationAdjustedScore", ascending=False)

    # 3. 현대적 버핏 (ROE_5Y_Avg 포함)
    mask_modern = df.apply(lambda r: enhanced_buffett_modern_filter(r, CONFIG), axis=1)
    raw_modern = df[mask_modern].copy()
    if not raw_modern.empty:
        scored_modern = build_modern_buffett_scores(raw_modern, CONFIG)
        scored_modern = scored_modern[scored_modern['TotalScore_Modern'] >= 70]

        # ROE_5Y_Avg 데이터 정리
        if 'ROE_5Y_Avg' not in scored_modern.columns and 'ROE(info)' in scored_modern.columns:
            scored_modern['ROE_5Y_Avg'] = scored_modern['ROE(info)']

        # 버핏 결과 컬럼 정리 및 순서 적용
        scored_modern = clean_buffett_columns(scored_modern, "modern_buffett")
        results["modern_buffett"] = scored_modern.sort_values("TotalScore_Modern", ascending=False)

    # 4. 트레이딩 프로파일 (기존대로 유지)
    for prof in ("swing", "daytrade"):
        mask_tr = df.apply(lambda r: pass_trading(r, prof, CONFIG), axis=1)
        base = df[mask_tr].copy()
        if not base.empty:
            scored = build_scores_trading(base, profile=prof, cfg=CONFIG)
            trading_cols = [
                "Ticker", "Name", "Sector", "Price", "DollarVol($M)", "RVOL",
                "ATR_PCT", "SMA20", "SMA50", "RET5", "RET20",
                "MomentumScore", "TrendScore", "LiquidityScore", "VolatilityScore", "TotalScore"
            ]
            trading_cols = [c for c in trading_cols if c in scored.columns]
            results[prof] = scored[trading_cols].sort_values("TotalScore", ascending=False)

    # 데이터 품질 리포트 출력
    print("\n=== 데이터 품질 리포트 ===")
    check_data_quality_issues(df)

    # ROE_5Y_Avg 데이터 품질 확인
    for profile_name, result_df in results.items():
        if not result_df.empty and 'ROE_5Y_Avg' in result_df.columns:
            non_null_count = result_df['ROE_5Y_Avg'].notna().sum()
            print(f"   {profile_name}: ROE_5Y_Avg {non_null_count}/{len(result_df)} 개 데이터")

    # 모든 설명이 포함된 엑셀 파일 생성
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_name = f"버핏장타&단타_종목_정리_{ts}.xlsx"

    print("\n🎨 엑셀 스타일링 및 상세 설명 추가 중...")
    create_styled_excel_output(results, out_name)

    print(f"\n🎯 COMPREHENSIVE SCREENER 완료: {out_name}")
    print("📚 포함된 설명 시트:")
    print("   - 버핏_지표_설명: 41개 버핏 프로파일 지표 상세 설명")
    print("   - 트레이딩_지표_설명: 16개 트레이딩 지표 상세 설명")
    print("   - 프로파일_비교: 5개 프로파일 특징 비교")
    print("   - ROE_비교_설명: ROE(info) vs ROE_5Y_Avg 비교 설명")

    return results


def check_data_quality_issues(df):
    """
    데이터 품질 문제 확인
    """
    # 문제가 될 수 있는 컬럼들
    problematic_columns = ['Debt_to_Equity', 'BuybackYield', 'P_FFO', 'FCF_Yield', 'PEG', 'EV_EBITDA']

    print("📊 데이터 가용성 현황:")
    for col in problematic_columns:
        if col in df.columns:
            non_null_count = df[col].notna().sum()
            total_count = len(df)
            percentage = (non_null_count / total_count) * 100 if total_count > 0 else 0

            status = "✅ 양호" if percentage > 50 else "⚠️ 부족" if percentage > 10 else "❌ 심각"
            print(f"   {col}: {non_null_count}/{total_count} ({percentage:.1f}%) - {status}")

    # Debt_to_Equity, BuybackYield, P_FFO가 비어있는 이유 설명
    print("\n💡 데이터 부족 이유:")
    print("   - Debt_to_Equity: yfinance에서 부채/자본비율 데이터 제공 불완전")
    print("   - BuybackYield: 자사주매수 수익률 데이터는 대부분 제공되지 않음")
    print("   - P_FFO: 리츠(REITs) 전용 지표로 일반 주식에는 적용되지 않음")
    print("   - FCF_Yield, PEG, EV_EBITDA: 계산에 필요한 기초 데이터 부족")


def _winsor_series(s: pd.Series, p=0.02):
    s = s.astype(float)
    lo, hi = s.quantile(p), s.quantile(1 - p)
    return s.clip(lower=lo, upper=hi)


def _percentile_rank(s: pd.Series, higher=True):
    s = s.astype(float)
    if not higher:
        s = -s
    return s.rank(pct=True, method="average")


def _clip01(x):
    try:
        return max(0.0, min(1.0, float(x)))
    except Exception:
        return np.nan


def check_data_quality_before_screening(df):
    """스크리너 실행 전 데이터 품질 확인"""
    print("=== 데이터 품질 확인 ===")

    essential_columns = {
        '버핏 분석': ['Price', 'MktCap($B)', 'RevYoY', 'OpMarginTTM', 'ROE(info)', 'PE', 'EV_EBITDA'],
        '트레이딩 분석': ['SMA20', 'SMA50', 'ATR_PCT', 'RVOL', 'RET5', 'RET20']
    }

    for category, columns in essential_columns.items():
        print(f"\n{category}:")
        for col in columns:
            if col in df.columns:
                non_null = df[col].notna().sum()
                pct = (non_null / len(df)) * 100
                print(f"  {col}: {non_null}/{len(df)} ({pct:.1f}%)")
            else:
                print(f"  {col}: ❌ 컬럼 없음")

    # NULL 비율이 높은 컬럼 식별
    low_quality_cols = []
    for col in df.columns:
        if df[col].notna().sum() / len(df) < 0.3:  # 30% 미만 데이터
            low_quality_cols.append(col)

    if low_quality_cols:
        print(f"\n⚠️ 주의: 데이터가 부족한 컬럼들: {low_quality_cols}")


def apply_winsorization_to_whole_dataset(df):
    """
    전체 데이터셋에 대해 Winsorization 적용 (일관성 보장)
    """
    df_processed = df.copy()

    # Winsorization 적용할 컬럼들
    winsorize_columns = [
        "RevYoY", "OpMarginTTM", "OperatingMargins(info)", "ROE(info)", "ROE_5Y_Avg",
        "FCF_Yield", "EV_EBITDA", "PE", "PEG", "PB", "DivYield", "Debt_to_Equity",
        "DollarVol($M)", "RVOL", "ATR_PCT", "RET5", "RET20"
    ]

    for col in winsorize_columns:
        if col in df_processed.columns and df_processed[col].notna().sum() > 0:
            df_processed[col] = _winsor_series(df_processed[col].astype(float), p=0.02)

    return df_processed


def enhanced_build_scores_buffett(df: pd.DataFrame, cfg=CONFIG):
    """기술적 지표를 반영한 개선된 버핏 점수 계산"""
    temp = df.copy()

    # 기존 점수 계산
    temp = build_scores_buffett(temp, cfg)

    # 기술적 지표 점수 추가
    tech_scores = []
    for idx, row in temp.iterrows():
        tech_score = calculate_technical_score(row, "buffett")
        tech_scores.append(tech_score)

    temp["TechnicalScore"] = pd.Series(tech_scores, index=temp.index)

    # 종합 점수에 기술적 지표 반영 (10% 가중치)
    temp["TotalScore_Enhanced"] = (
            temp["TotalScore"] * 0.9 +
            temp["TechnicalScore"] * 100 * 0.1
    )

    return temp


def build_scores_buffett(df: pd.DataFrame, cfg=CONFIG):
    """개선된 버핏 스타일 점수 계산 (데이터 누락 대응)"""
    temp = df.copy()

    # 누락될 수 있는 컬럼들에 대한 안전장치
    if "ROE_5Y_Avg" not in temp.columns:
        temp["ROE_5Y_Avg"] = temp["ROE(info)"]  # 기본값으로 ROE(info) 사용

    if "Debt_to_Equity" not in temp.columns:
        temp["Debt_to_Equity"] = np.nan

    if "BuybackYield" not in temp.columns:
        temp["BuybackYield"] = np.nan

    if "P_FFO" not in temp.columns:
        temp["P_FFO"] = np.nan

    if "FCF_Yield" not in temp.columns:
        temp["FCF_Yield"] = np.nan

    # 데이터 전처리
    temp["_OpMarginUse"] = temp[["OpMarginTTM", "OperatingMargins(info)"]].max(axis=1, numeric_only=True)

    growth_s = [];
    qual_s = [];
    val_s = [];
    cat_s = []

    for i, row in temp.iterrows():
        sec = str(row.get("Sector") or "").lower()

        # 성장 점수: 매출성장 + EPS 성장 기대
        rev_growth = row.get("RevYoY") or 0
        # PEG가 낮을수록 성장성 좋음 (역수 사용, 데이터 있을 때만)
        peg = row.get("PEG")
        if peg and not pd.isna(peg) and peg > 0:
            peg_score = 1.0 / peg
        else:
            peg_score = 0  # 데이터 없으면 0
        growth_components = [rev_growth, peg_score]
        growth_components = [x for x in growth_components if not pd.isna(x)]
        growth_score = np.nanmean(growth_components) if growth_components else 0
        growth_s.append(growth_score)

        # 질 점수: 수익성 + 재무건전성
        quality_components = []

        # 수익성 지표
        op_margin = row.get("_OpMarginUse")
        roe = row.get("ROE(info)") or row.get("ROE_5Y_Avg")
        if op_margin and not pd.isna(op_margin):
            quality_components.append(op_margin)
        if roe and not pd.isna(roe):
            quality_components.append(roe)

        # 재무건전성 지표
        debt_equity = row.get("Debt_to_Equity")
        if debt_equity is not None and not pd.isna(debt_equity):
            # 부채비율이 낮을수록 점수 높음
            debt_score = max(0, 1.0 - (debt_equity / cfg.get("MAX_DEBT_EQUITY", 2.0)))
            quality_components.append(debt_score)

        # FCF Yield (현금창출능력)
        fcf_yield = row.get("FCF_Yield")
        if fcf_yield and not pd.isna(fcf_yield) and fcf_yield > 0:
            quality_components.append(fcf_yield)

        qual_s.append(np.nanmean(quality_components) if quality_components else 0.5)

        # 가치 점수: 섹터별 차별화
        val_components = []

        if any(x in sec for x in FIN_SECTORS):
            # 금융주: P/B, ROE, Div Yield
            if "PB" in temp.columns and not pd.isna(row.get("PB")):
                val_components.append(_percentile_rank(temp["PB"], False)[i])
            if "ROE(info)" in temp.columns and not pd.isna(row.get("ROE(info)")):
                val_components.append(_percentile_rank(temp["ROE(info)"], True)[i])
            if "DivYield" in temp.columns and not pd.isna(row.get("DivYield")):
                val_components.append(_percentile_rank(temp["DivYield"], True)[i])

        elif any(x in sec for x in REIT_SECTORS):
            # 리츠: P/FFO, Div Yield
            if "P_FFO" in temp.columns and not pd.isna(row.get("P_FFO")):
                val_components.append(_percentile_rank(temp["P_FFO"], False)[i])
            if "DivYield" in temp.columns and not pd.isna(row.get("DivYield")):
                val_components.append(_percentile_rank(temp["DivYield"], True)[i])
        else:
            # 일반 주식: 다양한 가치 지표 (데이터 있는 것만 사용)
            for col, higher in [("FCF_Yield", True), ("EV_EBITDA", False),
                                ("PE", False), ("PEG", False), ("PB", False)]:
                if col in temp.columns and not pd.isna(row.get(col)):
                    val_components.append(_percentile_rank(temp[col], higher)[i])

        # val_components가 비어있지 않을 때만 계산
        if val_components:
            val_score = np.nanmean(val_components)
        else:
            val_score = 0.5  # 기본값

        val_s.append(val_score)

        # 촉매 점수: 배당, 자사주 매입
        catalyst_components = []
        if "DivYield" in temp and not pd.isna(row.get("DivYield")):
            catalyst_components.append(_percentile_rank(temp["DivYield"], True)[i])

        if "BuybackYield" in temp and not pd.isna(row.get("BuybackYield")):
            catalyst_components.append(_percentile_rank(temp["BuybackYield"], True)[i])

        # 거래량/변동성 (주목도 지표)
        if "RVOL" in temp and not pd.isna(row.get("RVOL")):
            rvol_score = min(1.0, (row.get("RVOL") or 1) / 3.0)
            catalyst_components.append(rvol_score)

        cat_s.append(np.nanmean(catalyst_components) if catalyst_components else 0.5)

    # 점수 정규화
    temp["GrowthScore"] = pd.Series(growth_s, index=temp.index).rank(pct=True).fillna(0.5)
    temp["QualityScore"] = pd.Series(qual_s, index=temp.index).rank(pct=True).fillna(0.5)
    temp["ValueScore"] = pd.Series(val_s, index=temp.index).fillna(0.5)
    temp["CatalystScore"] = pd.Series(cat_s, index=temp.index).fillna(0.5)

    # 총점 계산
    temp["TotalScore"] = 100 * (
            cfg.get("W_GROWTH", 0.15) * temp["GrowthScore"] +
            cfg.get("W_QUALITY", 0.35) * temp["QualityScore"] +
            cfg.get("W_VALUE", 0.40) * temp["ValueScore"] +
            cfg.get("W_CATALYST", 0.10) * temp["CatalystScore"]
    )

    return temp


def enhanced_build_scores_trading(df: pd.DataFrame, profile, cfg=CONFIG):
    """새로운 기술적 지표를 반영한 개선된 트레이딩 점수 계산"""
    temp = df.copy()

    # 데이터 전처리
    for col in ["RET5", "RET20", "RSI_14", "MACD", "MACD_Histogram", "BB_Position"]:
        if col in temp.columns:
            temp[col] = temp[col].astype(float).fillna(0)

    # 1. 모멘텀 점수 (기존 + 새로운 지표)
    momentum_components = []

    # 기존 모멘텀 지표
    if "RET5" in temp.columns and "RET20" in temp.columns:
        ret5_rank = temp["RET5"].rank(pct=True)
        ret20_rank = temp["RET20"].rank(pct=True)
        momentum_components.extend([ret5_rank, ret20_rank])

    # 새로운 기술적 모멘텀 지표
    if "MACD_Histogram" in temp.columns:
        macd_momentum = temp["MACD_Histogram"].rank(pct=True)
        momentum_components.append(macd_momentum)

    if "RSI_14" in temp.columns:
        # RSI가 50-70 사이면 모멘텀 좋음
        rsi_momentum = temp["RSI_14"].apply(lambda x: max(0, (x - 30) / 40) if 30 <= x <= 70 else 0.3)
        momentum_components.append(rsi_momentum)

    temp["MomentumScore"] = np.mean(momentum_components, axis=0) if momentum_components else 0.5

    # 2. 트렌드 점수 (개선)
    trend_components = []

    # 이동평균 트렌드
    close = temp["Price"]
    s20 = temp["SMA20"]
    s50 = temp["SMA50"]

    trend = []
    for i in temp.index:
        c, sma20, sma50 = close[i], s20[i], s50[i]
        score = 0.5
        try:
            if all(x is not None and not pd.isna(x) for x in [c, sma20, sma50]):
                if c > sma20 > sma50:
                    score = 1.0
                elif c > sma20:
                    score = 0.75
                elif sma20 and sma50 and sma20 > sma50:
                    score = 0.65
                else:
                    score = 0.25
        except Exception:
            score = 0.5
        trend.append(score)

    trend_components.append(pd.Series(trend, index=temp.index))

    # MACD 트렌드
    if "MACD" in temp.columns and "MACD_Signal" in temp.columns:
        macd_trend = (temp["MACD"] > temp["MACD_Signal"]).astype(float)
        trend_components.append(macd_trend)

    # 볼린저밴드 트렌드
    if "BB_Position" in temp.columns:
        bb_trend = temp["BB_Position"].apply(lambda x: max(0, min(1, x)))
        trend_components.append(bb_trend)

    temp["TrendScore"] = np.mean(trend_components, axis=0) if trend_components else 0.5

    # 3. 유동성 점수 (기존과 동일)
    dl = temp["DollarVol($M)"].rank(pct=True) if "DollarVol($M)" in temp.columns else pd.Series(0.5, index=temp.index)
    rv = temp["RVOL"].fillna(1.0).rank(pct=True) if "RVOL" in temp.columns else pd.Series(0.5, index=temp.index)
    temp["LiquidityScore"] = np.mean([dl, rv], axis=0)

    # 4. 변동성 점수 (개선)
    flt = cfg["SWING_FILTERS"] if profile == "swing" else cfg["DAY_FILTERS"]
    lo, hi = flt["ATR_PCT_RANGE"]
    target = (lo + hi) / 2.0
    sigma = (hi - lo) / 2.0

    vols = []
    for v in temp["ATR_PCT"].fillna(target):
        try:
            s = math.exp(-((float(v) - target) ** 2) / (2 * (sigma ** 2)))
        except Exception:
            s = 0.5
        vols.append(s)

    temp["VolatilityScore"] = pd.Series([max(0, min(1, x)) for x in vols], index=temp.index)

    # 5. 기술적 지표 종합 점수 (신규)
    tech_scores = []
    for idx, row in temp.iterrows():
        tech_score = calculate_technical_score(row, "trading")
        tech_scores.append(tech_score)

    temp["TechnicalScore"] = pd.Series(tech_scores, index=temp.index)

    # 종합 점수 계산 (새로운 가중치 적용)
    weights = cfg["TRADING_WEIGHTS"][profile]
    temp["TotalScore"] = 100 * (
            weights["momentum"] * temp["MomentumScore"] +
            weights["trend"] * temp["TrendScore"] +
            weights["liquidity"] * temp["LiquidityScore"] +
            weights["volatility"] * temp["VolatilityScore"] +
            weights["technical"] * temp["TechnicalScore"]
    )

    return temp


def build_scores_trading(df: pd.DataFrame, profile, cfg=CONFIG):
    temp = df.copy()
    for col in ["RET5", "RET20"]:
        if col in temp.columns:
            temp[col] = _winsor_series(temp[col].astype(float).fillna(0), p=0.02)
        else:
            temp[col] = 0.0
    mom = np.nanmean([_percentile_rank(temp["RET5"], True),
                      _percentile_rank(temp["RET20"], True)], axis=0)
    temp["MomentumScore"] = pd.Series(mom, index=temp.index).fillna(0.5)

    dl = _percentile_rank(temp["DollarVol($M)"], True) if "DollarVol($M)" in temp.columns else pd.Series(0.5,
                                                                                                         index=temp.index)
    rv = _percentile_rank(temp["RVOL"].fillna(1.0), True) if "RVOL" in temp.columns else pd.Series(0.5,
                                                                                                   index=temp.index)
    temp["LiquidityScore"] = np.nanmean([dl, rv], axis=0)

    close = temp["Price"];
    s20 = temp["SMA20"];
    s50 = temp["SMA50"]
    trend = []
    for i in temp.index:
        c, sma20, sma50 = close[i], s20[i], s50[i]
        score = 0.5
        try:
            if (c is not None) and (sma20 is not None) and (sma50 is not None):
                if c > sma20 > sma50:
                    score = 1.0
                elif c > sma20:
                    score = 0.75
                elif sma20 and sma50 and sma20 > sma50:
                    score = 0.65
                else:
                    score = 0.25
        except Exception:
            score = 0.5
        trend.append(score)
    temp["TrendScore"] = pd.Series([_clip01(x) for x in trend], index=temp.index)

    flt = cfg["SWING_FILTERS"] if profile == "swing" else cfg["DAY_FILTERS"]
    lo, hi = flt["ATR_PCT_RANGE"];
    target = (lo + hi) / 2.0;
    sigma = (hi - lo) / 2.0
    vols = []
    for v in temp["ATR_PCT"].fillna(target):
        try:
            s = math.exp(-((float(v) - target) ** 2) / (2 * (sigma ** 2)))
        except Exception:
            s = 0.5
        vols.append(s)
    temp["VolatilityScore"] = pd.Series([_clip01(x) for x in vols], index=temp.index)

    weights = {"swing": {"momentum": 0.45, "trend": 0.25, "liquidity": 0.20, "volatility": 0.10},
               "daytrade": {"momentum": 0.30, "trend": 0.10, "liquidity": 0.40, "volatility": 0.20}}[profile]
    temp["TotalScore"] = 100 * (weights["momentum"] * temp["MomentumScore"]
                                + weights["trend"] * temp["TrendScore"]
                                + weights["liquidity"] * temp["LiquidityScore"]
                                + weights["volatility"] * temp["VolatilityScore"])
    return temp


def load_cache(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Details cache not found: {path}")
    df = pd.read_csv(path)
    # 타입 보정
    num_cols = ["Price", "DollarVol($M)", "SMA20", "SMA50", "ATR_PCT", "RVOL", "RET5", "RET20",
                "MktCap($B)", "RevYoY", "OpMarginTTM", "OperatingMargins(info)", "ROE(info)", "EV_EBITDA",
                "PE", "PEG", "FCF_Yield", "PB", "DivYield", "P_FFO", "BuybackYield"]
    for c in num_cols:
        if c in df.columns: df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def enhanced_technical_screener():
    """
    기술적 지표가 강화된 통합 스크리너
    """
    df = load_cache(CONFIG["DETAILS_CACHE_FILE"])
    df = apply_winsorization_to_whole_dataset(df)

    print("🔍 기술적 지표 기반 스크리닝 시작...")

    # 적정가 계산
    fair_values_df = calculate_enhanced_fair_value(df)
    df = pd.concat([df, fair_values_df], axis=1)

    results = {}

    # 1. 기술적 지표 반영 버핏-Lite
    mask_lite = df.apply(lambda r: enhanced_pass_buffett_base(r, CONFIG), axis=1)
    raw_lite = df[mask_lite].copy()
    if not raw_lite.empty:
        scored_lite = enhanced_build_scores_buffett(raw_lite, CONFIG)
        scored_lite = scored_lite[scored_lite['TotalScore'] >= 60]
        scored_lite = clean_buffett_columns(scored_lite, 'buffett_lite')
        results["buffett_lite"] = scored_lite.sort_values("TotalScore", ascending=False)

    # 2. 기술적 지표 반영 버핏-Strict
    strict_cfg = CONFIG.copy()
    strict_cfg.update({
        "MIN_MKTCAP": 2_000_000_000,
        "MIN_PRICE": 10.0,
        "MIN_DOLLAR_VOLUME": 10_000_000,
        "MIN_DISCOUNT_PCT": 12.0,
        "MIN_OP_MARGIN_HF": 0.12,
        "MIN_REV_TTM_YOY_HF": 0.06,
        "HARD_PE_MAX": 20.0,
        "MIN_ROE_HF": 0.15,
        "MAX_DEBT_EQUITY": 1.0,
    })
    mask_strict = df.apply(lambda r: enhanced_pass_buffett_base(r, strict_cfg), axis=1)
    raw_strict = df[mask_strict].copy()
    if not raw_strict.empty:
        scored_strict = enhanced_build_scores_buffett(raw_strict, strict_cfg)
        scored_strict = scored_strict[scored_strict['TotalScore'] >= 70]
        scored_strict = clean_buffett_columns(scored_strict, 'buffett_strict')
        results["buffett_strict"] = scored_strict.sort_values("TotalScore", ascending=False)

    # 3. 현대적 버핏 (기술적 지표 반영)
    mask_modern = df.apply(lambda r: enhanced_buffett_modern_filter(r, CONFIG), axis=1)
    raw_modern = df[mask_modern].copy()
    if not raw_modern.empty:
        scored_modern = build_modern_buffett_scores(raw_modern, CONFIG)

        # 기술적 지표 점수 추가 반영
        tech_scores = []
        for idx, row in scored_modern.iterrows():
            tech_score = calculate_technical_score(row, "buffett")
            tech_scores.append(tech_score)

        scored_modern["TechnicalScore"] = pd.Series(tech_scores, index=scored_modern.index)
        scored_modern["TotalScore"] = (
                scored_modern["TotalScore"] * 0.9 +
                scored_modern["TechnicalScore"] * 100 * 0.1
        )

        scored_modern = scored_modern[scored_modern['TotalScore'] >= 70]
        scored_modern = clean_buffett_columns(scored_modern, 'modern_buffett')
        results["modern_buffett"] = scored_modern.sort_values("TotalScore", ascending=False)

    # 4. 기술적 지표 강화 트레이딩 프로파일
    for prof in ("swing", "daytrade"):
        mask_tr = df.apply(lambda r: enhanced_pass_trading(r, prof, CONFIG), axis=1)
        base = df[mask_tr].copy()
        if not base.empty:
            scored = enhanced_build_scores_trading(base, prof, CONFIG)

            # 트레이딩 출력 컬럼 (새로운 기술적 지표 포함)
            trading_cols = [
                "Ticker", "Name", "Sector", "Price", "DollarVol($M)", "RVOL",
                "ATR_PCT", "SMA20", "SMA50", "RET5", "RET20",
                "RSI_14", "MACD", "MACD_Histogram", "BB_Position", "High_52W_Ratio",
                "MomentumScore", "TrendScore", "LiquidityScore", "VolatilityScore",
                "TechnicalScore", "TotalScore"
            ]
            trading_cols = [c for c in trading_cols if c in scored.columns]
            results[f"{prof}"] = scored[trading_cols].sort_values("TotalScore", ascending=False)

    # 결과 출력
    print("\n=== 기술적 지표 기반 스크리닝 결과 ===")
    for profile_name, result_df in results.items():
        print(f"   {profile_name}: {len(result_df)}개 종목")

    # 엑셀 저장
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_name = f"기술적지표_강화_스크리너_{ts}.xlsx"

    with pd.ExcelWriter(out_name, engine='openpyxl') as writer:
        for profile_name, result_df in results.items():
            if not result_df.empty:
                result_df.to_excel(writer, sheet_name=profile_name[:31], index=False)

                # 향상된 조건부 서식 적용
                apply_enhanced_conditional_formatting(writer.sheets[profile_name[:31]],
                                                      result_df, profile_name)

        print(f"\n💾 결과 저장: {out_name}")

    print("\n🎨 엑셀 스타일링 및 상세 설명 추가 중...")
    create_styled_excel_output(results, out_name)

    print(f"\n🎯 COMPREHENSIVE SCREENER 완료: {out_name}")
    print("📚 포함된 설명 시트:")
    print("   - 버핏_지표_설명: 41개 버핏 프로파일 지표 상세 설명")
    print("   - 트레이딩_지표_설명: 16개 트레이딩 지표 상세 설명")
    print("   - 프로파일_비교: 5개 프로파일 특징 비교")
    print("   - ROE_비교_설명: ROE(info) vs ROE_5Y_Avg 비교 설명")


    return results


def enhanced_pass_trading(row, profile, cfg=CONFIG):
    """기술적 지표를 고려한 개선된 트레이딩 필터"""
    f = cfg["SWING_FILTERS"] if profile == "swing" else cfg["DAY_FILTERS"]

    # 기본 필터
    price = row.get("Price")
    dv = (row.get("DollarVol($M)") or 0) * 1_000_000
    rvol = row.get("RVOL")
    atr = row.get("ATR_PCT")

    if price is None or dv is None:
        return False
    if price < f["MIN_PRICE"] or dv < f["MIN_DOLLAR_VOLUME"]:
        return False
    if (rvol is None) or (rvol < f["MIN_RVOL"]):
        return False

    lo, hi = f["ATR_PCT_RANGE"]
    if (atr is None) or (atr < lo) or (atr > hi):
        return False

    # 기술적 지표 필터
    rsi = row.get("RSI_14")
    if rsi is not None:
        rsi_range = f.get("RSI_RANGE", [30, 70])
        if not (rsi_range[0] <= rsi <= rsi_range[1]):
            return False

    # MACD 조건
    macd_condition = f.get("MACD_CONDITION", "any")
    if macd_condition != "any":
        macd_histogram = row.get("MACD_Histogram")
        if macd_histogram is not None:
            if macd_condition == "positive" and macd_histogram <= 0:
                return False
            elif macd_condition == "negative" and macd_histogram >= 0:
                return False

    # 볼린저밴드 조건
    bb_condition = f.get("BB_CONDITION", "any")
    if bb_condition != "any":
        bb_position = row.get("BB_Position")
        if bb_position is not None:
            if bb_condition == "middle" and not (0.3 <= bb_position <= 0.7):
                return False
            elif bb_condition == "upper" and bb_position < 0.5:
                return False
            elif bb_condition == "lower" and bb_position > 0.5:
                return False

    # 52주 고가 비율
    min_52w_ratio = f.get("MIN_52W_RATIO", 0.7)
    high_52w_ratio = row.get("High_52W_Ratio")
    if high_52w_ratio is not None and high_52w_ratio < min_52w_ratio:
        return False

    # 기존 트렌드 필터
    rule = f.get("TREND_RULE", "any").lower()
    sma20 = row.get("SMA20")
    sma50 = row.get("SMA50")

    if rule == "close>sma20>sma50":
        if not (price and sma20 and sma50 and (price > sma20 > sma50)):
            return False
    elif rule == "sma20>50":
        if not (sma20 and sma50 and sma20 > sma50):
            return False

    if profile == "swing":
        ret20 = row.get("RET20")
        if ret20 is not None and ret20 < f["MIN_RET20"]:
            return False

    if profile == "daytrade":
        ret5 = row.get("RET5")
        if ret5 is not None and ret5 < f["MIN_RET5"]:
            return False

    return True


def pass_trading(row, profile, cfg=CONFIG):
    f = cfg["SWING_FILTERS"] if profile == "swing" else cfg["DAY_FILTERS"]
    price = row.get("Price");
    dv = (row.get("DollarVol($M)") or 0) * 1_000_000
    rvol = row.get("RVOL");
    atr = row.get("ATR_PCT")
    if price is None or dv is None: return False
    if price < f["MIN_PRICE"] or dv < f["MIN_DOLLAR_VOLUME"]: return False
    if (rvol is None) or (rvol < f["MIN_RVOL"]): return False
    lo, hi = f["ATR_PCT_RANGE"]
    if (atr is None) or (atr < lo) or (atr > hi): return False
    rule = f.get("TREND_RULE", "any").lower()
    sma20 = row.get("SMA20");
    sma50 = row.get("SMA50")
    if rule == "close>sma20>sma50":
        if not (price and sma20 and sma50 and (price > sma20 > sma50)): return False
    elif rule == "sma20>50":
        if not (sma20 and sma50 and sma20 > sma50): return False
    if profile == "swing":
        ret20 = row.get("RET20")
        if ret20 is not None and ret20 < f["MIN_RET20"]: return False
    if profile == "daytrade":
        ret5 = row.get("RET5")
        if ret5 is not None and ret5 < f["MIN_RET5"]: return False
    return True


if __name__ == "__main__":
    print("🚀 기술적 지표 강화 스크리너 실행 중...")
    results = enhanced_technical_screener()

    print("\n🎯 완료! 주요 특징:")
    print("   • RSI, MACD, 볼린저밴드 등 기술적 지표 반영")
    print("   • 버핏형 투자에 기술적 타이밍 요소 추가")
    print("   • 트레이딩 점수 계산에 다양한 기술적 지표 통합")
    print("   • 새로운 조건부 서식으로 시각적 분석 강화")