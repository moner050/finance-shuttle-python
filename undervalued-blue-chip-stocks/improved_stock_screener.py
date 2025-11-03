# -*- coding: utf-8 -*-
"""
improved_stock_screener.py
개선된 미국 주식 스크리너 - 저평가우량주, 장타, 단타 최적화

주요 개선사항:
1. 저평가우량주: PEG, FCF Yield, 배당수익률, EPS 성장률 추가
2. 장타 전략: 성장성 지표 강화, 52주 고가 비율 추가
3. 단타 전략: MACD, 볼린저밴드, 52주 고저가 비율 활용
4. 데이터 컬럼 정확성 개선
"""

import os
import math
import warnings
import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

warnings.filterwarnings("ignore", category=RuntimeWarning)


# ============================================================================
# 설정 클래스 (데이터클래스로 관리)
# ============================================================================

@dataclass
class FilterCriteria:
    """필터 기준 데이터클래스"""
    min_mktcap: float = 500_000_000  # 5억 달러 (중소형주 포함)
    min_price: float = 5.0  # 5달러 (페니스톡 제외)
    min_dollar_volume: float = 1_000_000  # 100만 달러 (유동성 기준 완화)
    max_pe: float = 35.0  # S&P500 평균 고려
    max_peg: float = 2.0  # PEG 비율 기준 추가
    min_rev_growth: float = -0.05  # -5% (턴어라운드 기회 포함)
    min_eps_growth: float = 0.0  # EPS 성장률 기준
    min_op_margin: float = 0.05  # 5% (섹터별 차이 고려)
    min_roe: float = 0.08  # 8% (현실적 기준)
    min_fcf_yield: float = 0.0  # FCF Yield 기준
    min_div_yield: float = 0.0  # 배당수익률 기준


class ScreenerConfig:
    """스크리너 설정 관리"""

    # 프로파일별 필터 기준
    PROFILES = {
        # 저평가 우량주 (워렌 버핏 스타일)
        'undervalued_quality': FilterCriteria(
            min_mktcap=2_000_000_000,  # 20억 달러 이상
            min_price=10.0,
            min_dollar_volume=5_000_000,  # 500만 달러
            max_pe=25.0,  # 합리적인 PER
            max_peg=1.5,  # PEG < 1.5
            min_rev_growth=0.05,  # 최소 5% 매출 성장
            min_eps_growth=0.05,  # 최소 5% EPS 성장
            min_op_margin=0.12,  # 12% 이상 영업이익률
            min_roe=0.15,  # 15% 이상 ROE
            min_fcf_yield=0.03,  # 3% 이상 FCF Yield
        ),

        # 기본 가치투자
        'value_basic': FilterCriteria(
            min_mktcap=500_000_000,
            max_pe=30.0,
            max_peg=2.0,
            min_op_margin=0.05,
            min_roe=0.08
        ),

        # 엄격한 가치투자
        'value_strict': FilterCriteria(
            min_mktcap=2_000_000_000,  # 20억 달러
            min_dollar_volume=5_000_000,  # 500만 달러
            max_pe=20.0,
            max_peg=1.5,
            min_rev_growth=0.05,
            min_eps_growth=0.05,
            min_op_margin=0.10,
            min_roe=0.12,
            min_fcf_yield=0.02
        ),

        # 성장+품질 (장타 전략)
        'growth_quality': FilterCriteria(
            min_mktcap=1_000_000_000,
            min_rev_growth=0.15,  # 15% 이상 매출 성장
            min_eps_growth=0.10,  # 10% 이상 EPS 성장
            min_op_margin=0.15,
            min_roe=0.15,
            max_pe=40.0,  # 성장주는 높은 PER 허용
            max_peg=2.0
        ),

        # 모멘텀 트레이딩 (단타)
        'momentum': {
            'min_price': 10.0,
            'min_volume': 3_000_000,
            'min_rvol': 1.3,  # 평균 대비 1.3배 이상 거래량
            'rsi_range': (40, 70),  # 과매도 영역 진입 후 반등
            'ret20_min': 0.03,  # 최근 20일 3% 이상 상승
            'high_52w_min': 0.7,  # 52주 고가의 70% 이상
            'macd_positive': True  # MACD 히스토그램 양수
        },

        # 스윙 트레이딩 (단타)
        'swing': {
            'min_price': 5.0,
            'min_volume': 1_000_000,
            'atr_range': (0.02, 0.10),  # 변동성 2-10%
            'rsi_range': (30, 70),
            'bb_position_range': (0.2, 0.8),  # 볼린저밴드 20-80% 위치
            'ret5_range': (-0.05, 0.10)  # 최근 5일 -5% ~ 10%
        }
    }

    # 점수 가중치 (프로파일별로 다르게 적용)
    SCORE_WEIGHTS = {
        'value': {'growth': 0.15, 'quality': 0.35, 'value': 0.40, 'momentum': 0.10},
        'growth': {'growth': 0.45, 'quality': 0.30, 'value': 0.15, 'momentum': 0.10},
        'balanced': {'growth': 0.25, 'quality': 0.30, 'value': 0.30, 'momentum': 0.15},
        'trading': {'growth': 0.05, 'quality': 0.15, 'value': 0.20, 'momentum': 0.60}
    }

    # 섹터별 조정 파라미터
    SECTOR_ADJUSTMENTS = {
        'technology': {'pe_multiplier': 1.4, 'margin_discount': 0.0, 'growth_premium': 1.2},
        'healthcare': {'pe_multiplier': 1.3, 'margin_discount': 0.1, 'growth_premium': 1.1},
        'financial': {'pe_multiplier': 0.8, 'margin_discount': 0.5, 'use_pb': True},
        'utilities': {'pe_multiplier': 0.9, 'margin_discount': 0.3, 'growth_premium': 0.8},
        'real estate': {'pe_multiplier': 1.0, 'margin_discount': 0.4, 'use_pb': True},
        'consumer': {'pe_multiplier': 1.1, 'margin_discount': 0.2, 'growth_premium': 1.0},
        'industrial': {'pe_multiplier': 1.0, 'margin_discount': 0.2, 'growth_premium': 0.9},
        'energy': {'pe_multiplier': 1.2, 'margin_discount': 0.3, 'growth_premium': 0.9},
    }


# ============================================================================
# 유틸리티 클래스
# ============================================================================

class DataProcessor:
    """데이터 처리 유틸리티"""

    @staticmethod
    def winsorize(series: pd.Series, limits: Tuple[float, float] = (0.01, 0.99)) -> pd.Series:
        """이상치 제거 (Winsorization)"""
        return series.clip(
            lower=series.quantile(limits[0]),
            upper=series.quantile(limits[1])
        )

    @staticmethod
    def normalize_score(series: pd.Series, ascending: bool = True) -> pd.Series:
        """점수 정규화 (0-1 범위)"""
        if not ascending:
            series = -series
        return series.rank(pct=True, method='average')

    @staticmethod
    def safe_divide(numerator: float, denominator: float, default: float = 0) -> float:
        """안전한 나눗셈"""
        try:
            if denominator and denominator != 0 and not pd.isna(denominator):
                return numerator / denominator
            return default
        except:
            return default

    @staticmethod
    def safe_value(value, default=0):
        """안전한 값 가져오기"""
        if pd.isna(value) or value is None:
            return default
        return value


# ============================================================================
# 가치평가 모델
# ============================================================================

class ValuationModel:
    """통합 가치평가 모델"""

    @staticmethod
    def calculate_fair_value(df: pd.DataFrame) -> pd.DataFrame:
        """적정가치 계산 (개선된 버전)"""
        fair_values = []

        for idx, row in df.iterrows():
            price = row.get('Price', 0)
            pe = row.get('PE', 0)
            pb = row.get('PB', 0)
            peg = row.get('PEG', 0)
            sector = str(row.get('Sector', '')).lower()

            # 섹터 평균 대비 상대가치
            sector_data = df[df['Sector'] == row['Sector']]

            valuations = []

            # 1. PE 기반 가치 (섹터 중앙값 사용)
            if pe > 0 and len(sector_data) > 3:
                sector_pe_median = sector_data['PE'][sector_data['PE'] > 0].median()
                if sector_pe_median and not pd.isna(sector_pe_median) and pe > 0:
                    eps = DataProcessor.safe_divide(price, pe)
                    pe_value = sector_pe_median * eps
                    if pe_value > 0:
                        valuations.append(pe_value)

            # 2. PB 기반 가치 (금융, 부동산)
            if pb > 0 and any(x in sector for x in ['financ', 'real', 'bank']):
                sector_pb_median = sector_data['PB'][sector_data['PB'] > 0].median()
                if sector_pb_median and not pd.isna(sector_pb_median) and pb > 0:
                    bps = DataProcessor.safe_divide(price, pb)
                    pb_value = sector_pb_median * bps
                    if pb_value > 0:
                        valuations.append(pb_value)

            # 3. PEG 기반 가치
            if peg > 0 and peg < 3:
                # PEG = 1일 때가 적정가치
                ideal_pe = DataProcessor.safe_divide(pe, peg, 0)
                if ideal_pe > 0 and pe > 0:
                    eps = DataProcessor.safe_divide(price, pe)
                    peg_value = ideal_pe * eps
                    if peg_value > 0:
                        valuations.append(peg_value)

            # 4. FCF 기반 가치
            fcf_yield = row.get('FCF_Yield', 0)
            if fcf_yield and fcf_yield > 0.02:  # 2% 이상
                # FCF Yield 역수를 배수로 사용
                fcf_multiple = DataProcessor.safe_divide(1, fcf_yield, 0)
                if fcf_multiple > 0:
                    mktcap = row.get('MktCap($B)', 0) * 1e9
                    if mktcap > 0:
                        fcf = mktcap * fcf_yield
                        median_fcf_yield = 0.05  # 중앙값 5% 가정
                        fcf_value = DataProcessor.safe_divide(fcf, median_fcf_yield, 0) / (row.get('MktCap($B)', 1) * 1e9) * price
                        if fcf_value > 0:
                            valuations.append(fcf_value)

            # 평균 적정가치
            if valuations:
                fair_value = np.median(valuations)  # 중앙값 사용 (이상치에 강건)
                discount = DataProcessor.safe_divide(fair_value - price, price, 0)
            else:
                fair_value = price
                discount = 0

            fair_values.append({
                'FairValue': fair_value,
                'Discount': discount
            })

        return pd.DataFrame(fair_values, index=df.index)


# ============================================================================
# 메인 스크리너 클래스
# ============================================================================

class StockScreener:
    """통합 주식 스크리너"""

    def __init__(self, config: ScreenerConfig = None):
        self.config = config or ScreenerConfig()
        self.processor = DataProcessor()
        self.valuation = ValuationModel()

    def load_data(self, filepath: str) -> pd.DataFrame:
        """데이터 로드 및 전처리"""
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"파일을 찾을 수 없습니다: {filepath}")

        df = pd.read_csv(filepath)

        # 숫자형 변환
        numeric_cols = [
            'Price', 'DollarVol($M)', 'MktCap($B)', 'PE', 'PEG', 'PB', 'PS',
            'ROE(info)', 'ROA(info)', 'OpMarginTTM', 'OperatingMargins(info)',
            'RevYoY', 'FCF_Yield', 'DivYield', 'PayoutRatio',
            'EPS_Growth_3Y', 'Revenue_Growth_3Y', 'EBITDA_Growth_3Y',
            'EV_EBITDA', 'Beta', 'ShortPercent', 'InsiderOwnership', 'InstitutionOwnership',
            'RVOL', 'RSI_14', 'RET5', 'RET20', 'RET63', 'ATR_PCT',
            'SMA20', 'SMA50', 'SMA200',
            'MACD', 'MACD_Signal', 'MACD_Histogram',
            'BB_Position', 'High_52W_Ratio', 'Low_52W_Ratio',
            'Momentum_12M', 'Volatility_21D'
        ]

        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        # 이상치 제거 (Winsorization)
        winsor_cols = ['PE', 'PB', 'PEG', 'PS', 'RevYoY', 'EV_EBITDA']
        for col in winsor_cols:
            if col in df.columns:
                df[col] = self.processor.winsorize(df[col])

        return df

    def apply_filters(self, df: pd.DataFrame, profile: str) -> pd.DataFrame:
        """프로파일별 필터 적용"""
        if profile not in self.config.PROFILES:
            raise ValueError(f"알 수 없는 프로파일: {profile}")

        if profile in ['undervalued_quality', 'value_basic', 'value_strict', 'growth_quality']:
            return self._apply_fundamental_filter(df, profile)
        else:
            return self._apply_trading_filter(df, profile)

    def _apply_fundamental_filter(self, df: pd.DataFrame, profile: str) -> pd.DataFrame:
        """펀더멘털 필터 (개선됨)"""
        criteria = self.config.PROFILES[profile]

        # 섹터별 조정 적용
        mask = pd.Series([True] * len(df), index=df.index)

        for idx, row in df.iterrows():
            sector = str(row.get('Sector', '')).lower()
            sector_key = next((k for k in self.config.SECTOR_ADJUSTMENTS if k in sector), None)
            sector_adj = self.config.SECTOR_ADJUSTMENTS.get(sector_key, {
                'pe_multiplier': 1.0,
                'margin_discount': 0,
                'growth_premium': 1.0
            })

            # 기본 필터
            if self.processor.safe_value(row.get('MktCap($B)', 0) * 1e9) < criteria.min_mktcap:
                mask[idx] = False
                continue

            if self.processor.safe_value(row.get('Price', 0)) < criteria.min_price:
                mask[idx] = False
                continue

            if self.processor.safe_value(row.get('DollarVol($M)', 0) * 1e6) < criteria.min_dollar_volume:
                mask[idx] = False
                continue

            # PE 필터 (섹터 조정)
            pe = self.processor.safe_value(row.get('PE', 0))
            if pe > 0:
                pe_limit = criteria.max_pe * sector_adj.get('pe_multiplier', 1.0)
                if pe > pe_limit:
                    mask[idx] = False
                    continue

            # PEG 필터
            peg = self.processor.safe_value(row.get('PEG', 0))
            if peg > 0 and peg > criteria.max_peg:
                mask[idx] = False
                continue

            # 매출 성장률 필터
            rev_yoy = self.processor.safe_value(row.get('RevYoY', 0))
            if rev_yoy < criteria.min_rev_growth:
                mask[idx] = False
                continue

            # EPS 성장률 필터
            eps_growth = self.processor.safe_value(row.get('EPS_Growth_3Y', 0))
            if eps_growth < criteria.min_eps_growth:
                mask[idx] = False
                continue

            # 마진 필터 (섹터 조정)
            margin_req = criteria.min_op_margin * (1 - sector_adj.get('margin_discount', 0))
            op_margin = self.processor.safe_value(row.get('OpMarginTTM')) or self.processor.safe_value(row.get('OperatingMargins(info)', 0))
            if op_margin < margin_req:
                mask[idx] = False
                continue

            # ROE 필터
            roe = self.processor.safe_value(row.get('ROE(info)', 0))
            if roe < criteria.min_roe:
                mask[idx] = False
                continue

            # FCF Yield 필터
            fcf_yield = self.processor.safe_value(row.get('FCF_Yield', 0))
            if fcf_yield < criteria.min_fcf_yield:
                mask[idx] = False
                continue

        return df[mask]

    def _apply_trading_filter(self, df: pd.DataFrame, profile: str) -> pd.DataFrame:
        """트레이딩 필터 (개선됨)"""
        criteria = self.config.PROFILES[profile]

        mask = (
            (df['Price'] >= criteria['min_price']) &
            (df['DollarVol($M)'] * 1e6 >= criteria['min_volume'])
        )

        if 'min_rvol' in criteria:
            mask &= (df['RVOL'] >= criteria['min_rvol'])

        if 'rsi_range' in criteria:
            rsi_min, rsi_max = criteria['rsi_range']
            mask &= (df['RSI_14'] >= rsi_min) & (df['RSI_14'] <= rsi_max)

        if 'atr_range' in criteria:
            atr_min, atr_max = criteria['atr_range']
            mask &= (df['ATR_PCT'] >= atr_min) & (df['ATR_PCT'] <= atr_max)

        if 'ret20_min' in criteria:
            mask &= (df['RET20'] >= criteria['ret20_min'])

        if 'ret5_range' in criteria:
            ret5_min, ret5_max = criteria['ret5_range']
            mask &= (df['RET5'] >= ret5_min) & (df['RET5'] <= ret5_max)

        if 'high_52w_min' in criteria:
            mask &= (df['High_52W_Ratio'] >= criteria['high_52w_min'])

        if 'bb_position_range' in criteria:
            bb_min, bb_max = criteria['bb_position_range']
            mask &= (df['BB_Position'] >= bb_min) & (df['BB_Position'] <= bb_max)

        if 'macd_positive' in criteria and criteria['macd_positive']:
            mask &= (df['MACD_Histogram'] > 0)

        return df[mask]

    def calculate_scores(self, df: pd.DataFrame, score_type: str = 'balanced') -> pd.DataFrame:
        """종합 점수 계산 (개선됨)"""
        weights = self.config.SCORE_WEIGHTS[score_type]

        # 성장 점수 (더 많은 지표 활용)
        growth_components = []
        if 'RevYoY' in df.columns:
            growth_components.append(self.processor.normalize_score(df['RevYoY'].fillna(0)))
        if 'EPS_Growth_3Y' in df.columns:
            growth_components.append(self.processor.normalize_score(df['EPS_Growth_3Y'].fillna(0)))
        if 'Revenue_Growth_3Y' in df.columns:
            growth_components.append(self.processor.normalize_score(df['Revenue_Growth_3Y'].fillna(0)))
        if 'RET20' in df.columns:
            growth_components.append(self.processor.normalize_score(df['RET20'].fillna(0)))

        growth_score = np.mean(growth_components, axis=0) if growth_components else 0.5

        # 품질 점수
        quality_components = []
        if 'ROE(info)' in df.columns:
            quality_components.append(self.processor.normalize_score(df['ROE(info)'].fillna(0)))
        if 'OpMarginTTM' in df.columns:
            quality_components.append(self.processor.normalize_score(df['OpMarginTTM'].fillna(0)))
        if 'FCF_Yield' in df.columns:
            quality_components.append(self.processor.normalize_score(df['FCF_Yield'].fillna(0)))
        if 'ROA(info)' in df.columns:
            quality_components.append(self.processor.normalize_score(df['ROA(info)'].fillna(0)))

        quality_score = np.mean(quality_components, axis=0) if quality_components else 0.5

        # 가치 점수
        value_components = []
        if 'PE' in df.columns:
            value_components.append(self.processor.normalize_score(df['PE'].fillna(100), ascending=False))
        if 'PEG' in df.columns:
            value_components.append(self.processor.normalize_score(df['PEG'].fillna(10), ascending=False))
        if 'PB' in df.columns:
            value_components.append(self.processor.normalize_score(df['PB'].fillna(10), ascending=False))
        if 'Discount' in df.columns:
            value_components.append(self.processor.normalize_score(df['Discount'].fillna(-1)))

        value_score = np.mean(value_components, axis=0) if value_components else 0.5

        # 모멘텀 점수 (더 많은 지표 활용)
        momentum_components = []
        if 'RVOL' in df.columns:
            momentum_components.append(self.processor.normalize_score(df['RVOL'].fillna(1)))
        if 'RSI_14' in df.columns:
            rsi_norm = (df['RSI_14'].fillna(50) - 30) / 40  # 30-70 범위 정규화
            momentum_components.append(rsi_norm.clip(0, 1))
        if 'RET5' in df.columns:
            momentum_components.append(self.processor.normalize_score(df['RET5'].fillna(0)))
        if 'High_52W_Ratio' in df.columns:
            momentum_components.append(self.processor.normalize_score(df['High_52W_Ratio'].fillna(0.5)))
        if 'MACD_Histogram' in df.columns:
            macd_norm = df['MACD_Histogram'].fillna(0).apply(lambda x: 1 if x > 0 else 0)
            momentum_components.append(macd_norm)

        momentum_score = np.mean(momentum_components, axis=0) if momentum_components else 0.5

        # 점수 저장
        df['GrowthScore'] = growth_score
        df['QualityScore'] = quality_score
        df['ValueScore'] = value_score
        df['MomentumScore'] = momentum_score

        df['TotalScore'] = (
            weights['growth'] * df['GrowthScore'] +
            weights['quality'] * df['QualityScore'] +
            weights['value'] * df['ValueScore'] +
            weights['momentum'] * df['MomentumScore']
        ) * 100

        return df

    def screen_stocks(self, filepath: str, min_score: float = 60) -> Dict[str, pd.DataFrame]:
        """전체 스크리닝 실행"""
        print("📊 데이터 로딩...")
        df = self.load_data(filepath)
        print(f"✅ {len(df)}개 종목 로드 완료")

        # 적정가치 계산
        print("💰 적정가치 계산 중...")
        fair_values = self.valuation.calculate_fair_value(df)
        df = pd.concat([df, fair_values], axis=1)

        results = {}

        # 프로파일별 스크리닝
        profiles = [
            ('undervalued_quality', 'value', 70),  # 저평가 우량주
            ('value_basic', 'value', 55),  # 기본 가치투자
            ('value_strict', 'value', 65),  # 엄격한 가치투자
            ('growth_quality', 'growth', 65),  # 성장+품질 (장타)
            ('momentum', 'trading', 65),  # 모멘텀 (단타)
            ('swing', 'trading', 60)  # 스윙 (단타)
        ]

        for profile_name, score_type, min_threshold in profiles:
            print(f"\n🔍 {profile_name} 스크리닝...")

            # 필터 적용
            try:
                filtered = self.apply_filters(df.copy(), profile_name)
            except Exception as e:
                print(f"   ⚠️ 필터 적용 오류: {e}")
                continue

            if filtered.empty:
                print(f"   ⚠️ 조건 충족 종목 없음")
                continue

            # 점수 계산
            try:
                scored = self.calculate_scores(filtered.copy(), score_type)
            except Exception as e:
                print(f"   ⚠️ 점수 계산 오류: {e}")
                continue

            # 최소 점수 필터
            final = scored[scored['TotalScore'] >= min_threshold]

            # 결과 정리
            if not final.empty:
                # 핵심 컬럼만 선택
                cols = self._select_columns(profile_name, final.columns)
                results[profile_name] = final[cols].sort_values('TotalScore', ascending=False)
                print(f"   ✅ {len(results[profile_name])}개 종목 발굴")
            else:
                print(f"   ⚠️ 최소 점수 충족 종목 없음")

        return results

    def _select_columns(self, profile: str, available_cols: List[str]) -> List[str]:
        """프로파일별 출력 컬럼 선택"""
        base_cols = ['Ticker', 'Name', 'Sector', 'Industry', 'Price', 'MktCap($B)']

        if profile in ['undervalued_quality', 'value_basic', 'value_strict', 'growth_quality']:
            specific_cols = [
                'FairValue', 'Discount', 'PE', 'PEG', 'PB', 'PS',
                'ROE(info)', 'OpMarginTTM', 'RevYoY', 'EPS_Growth_3Y', 'Revenue_Growth_3Y',
                'FCF_Yield', 'DivYield', 'EV_EBITDA',
                'Beta', 'InsiderOwnership', 'InstitutionOwnership',
                'GrowthScore', 'QualityScore', 'ValueScore', 'TotalScore'
            ]
        else:  # 트레이딩 전략
            specific_cols = [
                'DollarVol($M)', 'RVOL', 'ATR_PCT', 'Volatility_21D',
                'RSI_14', 'MACD', 'MACD_Histogram', 'BB_Position',
                'RET5', 'RET20', 'High_52W_Ratio', 'Low_52W_Ratio',
                'SMA20', 'SMA50', 'SMA200',
                'MomentumScore', 'TotalScore'
            ]

        return base_cols + [col for col in specific_cols if col in available_cols]


# ============================================================================
# 엑셀 출력 클래스
# ============================================================================

class ExcelExporter:
    """엑셀 출력 관리"""

    # 컬럼 설명 딕셔너리
    COLUMN_DESCRIPTIONS = {
        'Ticker': '티커 심볼',
        'Name': '회사명',
        'Sector': '섹터',
        'Industry': '산업군',
        'Price': '현재 주가',
        'MktCap($B)': '시가총액 (10억 달러)',
        'DollarVol($M)': '일평균 거래대금 (백만 달러)',

        # 가치평가
        'FairValue': '적정가치 (PE, PB, PEG, FCF 기반 계산)',
        'Discount': '할인율 (적정가치 대비 현재가 할인 정도)',
        'PE': 'PER (주가수익비율) - 낮을수록 저평가',
        'PEG': 'PEG 비율 (PER/성장률) - 1 이하 매력적',
        'PB': 'PBR (주가순자산비율) - 낮을수록 저평가',
        'PS': 'PSR (주가매출비율) - 낮을수록 저평가',
        'EV_EBITDA': 'EV/EBITDA 배수',

        # 수익성 지표
        'ROE(info)': '자기자본이익률 - 높을수록 우수',
        'ROA(info)': '총자산이익률 - 높을수록 우수',
        'OpMarginTTM': '영업이익률 (TTM) - 높을수록 우수',
        'OperatingMargins(info)': '영업이익률 (info)',

        # 성장성 지표
        'RevYoY': '매출 YoY 성장률',
        'EPS_Growth_3Y': '3년 EPS 성장률 (CAGR)',
        'Revenue_Growth_3Y': '3년 매출 성장률 (CAGR)',
        'EBITDA_Growth_3Y': '3년 EBITDA 성장률',

        # 현금흐름
        'FCF_Yield': 'FCF 수익률 (현금 창출 능력)',
        'DivYield': '배당수익률',
        'PayoutRatio': '배당성향',

        # 재무안정성
        'Beta': '베타 (시장 대비 변동성)',
        'ShortPercent': '공매도 비율',
        'InsiderOwnership': '내부자 지분율',
        'InstitutionOwnership': '기관 투자자 지분율',

        # 기술적 지표 (단타/장타)
        'RVOL': '상대 거래량 (평균 대비)',
        'RSI_14': 'RSI 14일 (30 이하 과매도, 70 이상 과매수)',
        'ATR_PCT': 'ATR 퍼센트 (변동성)',
        'Volatility_21D': '21일 변동성',
        'RET5': '5일 수익률',
        'RET20': '20일 수익률',
        'RET63': '3개월 수익률',
        'SMA20': '20일 이동평균',
        'SMA50': '50일 이동평균',
        'SMA200': '200일 이동평균',
        'MACD': 'MACD 선',
        'MACD_Signal': 'MACD 시그널 선',
        'MACD_Histogram': 'MACD 히스토그램 (양수 = 상승 추세)',
        'BB_Position': '볼린저밴드 위치 (0-1, 0.5 중앙)',
        'High_52W_Ratio': '52주 고가 대비 비율',
        'Low_52W_Ratio': '52주 저가 대비 비율',
        'Momentum_12M': '12개월 모멘텀',

        # 종합 점수
        'GrowthScore': '성장 점수 (0-100%)',
        'QualityScore': '품질 점수 (0-100%)',
        'ValueScore': '가치 점수 (0-100%)',
        'MomentumScore': '모멘텀 점수 (0-100%)',
        'TotalScore': '종합 점수 (0-100점)',
    }

    # 전략별 필터 기준
    STRATEGY_CRITERIA = {
        'undervalued_quality': {
            'name': '저평가 우량주 (워렌 버핏 스타일)',
            'criteria': [
                '시가총액: 20억 달러 이상',
                '주가: 10달러 이상',
                '거래대금: 500만 달러 이상',
                'PER < 25 (섹터별 조정)',
                'PEG < 1.5',
                '매출 성장률 > 5%',
                'EPS 성장률 > 5%',
                '영업이익률 > 12%',
                'ROE > 15%',
                'FCF Yield > 3%',
            ]
        },
        'value_basic': {
            'name': '기본 가치투자',
            'criteria': [
                '시가총액: 5억 달러 이상',
                '주가: 5달러 이상',
                '거래대금: 100만 달러 이상',
                'PER < 30 (섹터별 조정)',
                'PEG < 2.0',
                '영업이익률 > 5%',
                'ROE > 8%',
            ]
        },
        'value_strict': {
            'name': '엄격한 가치투자',
            'criteria': [
                '시가총액: 20억 달러 이상',
                '주가: 5달러 이상',
                '거래대금: 500만 달러 이상',
                'PER < 20 (섹터별 조정)',
                'PEG < 1.5',
                '매출 성장률 > 5%',
                'EPS 성장률 > 5%',
                '영업이익률 > 10%',
                'ROE > 12%',
                'FCF Yield > 2%',
            ]
        },
        'growth_quality': {
            'name': '성장+품질 (장타 전략)',
            'criteria': [
                '시가총액: 10억 달러 이상',
                '매출 성장률 > 15%',
                'EPS 성장률 > 10%',
                '영업이익률 > 15%',
                'ROE > 15%',
                'PER < 40 (성장주 특성 반영)',
                'PEG < 2.0',
            ]
        },
        'momentum': {
            'name': '모멘텀 트레이딩 (단타)',
            'criteria': [
                '주가: 10달러 이상',
                '거래대금: 300만 달러 이상',
                '상대 거래량 > 1.3배',
                'RSI: 40-70 (과매도 후 반등)',
                '20일 수익률 > 3%',
                '52주 고가 대비 > 70%',
                'MACD 히스토그램 > 0 (상승 추세)',
            ]
        },
        'swing': {
            'name': '스윙 트레이딩 (단타)',
            'criteria': [
                '주가: 5달러 이상',
                '거래대금: 100만 달러 이상',
                'ATR 변동성: 2-10%',
                'RSI: 30-70',
                '볼린저밴드 위치: 20-80%',
                '5일 수익률: -5% ~ 10%',
            ]
        },
    }

    @staticmethod
    def export(results: Dict[str, pd.DataFrame], filename: str = None):
        """결과를 엑셀로 출력"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            filename = f"stock_screener_{timestamp}.xlsx"

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Column Guide 시트 먼저 생성
            ExcelExporter._create_column_guide(writer)

            # 요약 시트
            summary_data = []
            for profile, df in results.items():
                if not df.empty:
                    top_tickers = ', '.join(df.head(5)['Ticker'].tolist())
                    avg_score = df['TotalScore'].mean()

                    # 추가 통계
                    avg_pe = df['PE'].mean() if 'PE' in df.columns else None
                    avg_growth = df['EPS_Growth_3Y'].mean() if 'EPS_Growth_3Y' in df.columns else None

                    summary_data.append({
                        'Profile': profile,
                        'Count': len(df),
                        'Avg Score': f"{avg_score:.1f}",
                        'Avg PE': f"{avg_pe:.1f}" if avg_pe else "N/A",
                        'Avg Growth': f"{avg_growth*100:.1f}%" if avg_growth else "N/A",
                        'Top 5 Tickers': top_tickers
                    })

            if summary_data:
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                ExcelExporter._format_sheet(writer, 'Summary', summary_df, is_summary=True)

            # 각 프로파일별 시트
            for profile, df in results.items():
                if not df.empty:
                    # 숫자 포맷 조정
                    df_copy = df.copy()

                    # 퍼센트 컬럼 (100 곱하고 % 추가)
                    pct_cols = ['Discount', 'ROE(info)', 'OpMarginTTM', 'RevYoY',
                                'EPS_Growth_3Y', 'Revenue_Growth_3Y', 'FCF_Yield', 'DivYield',
                                'ATR_PCT', 'RET5', 'RET20', 'Volatility_21D',
                                'GrowthScore', 'QualityScore', 'ValueScore', 'MomentumScore']

                    for col in pct_cols:
                        if col in df_copy.columns:
                            df_copy[col] = df_copy[col].apply(
                                lambda x: f"{x * 100:.2f}%" if pd.notna(x) else ""
                            )

                    # 소수점 2자리 컬럼
                    decimal_cols = ['FairValue', 'Price', 'TotalScore', 'PE', 'PEG', 'PB', 'PS',
                                   'RVOL', 'SMA20', 'SMA50', 'SMA200', 'RSI_14', 'MACD',
                                   'BB_Position', 'High_52W_Ratio', 'Low_52W_Ratio', 'Beta']

                    for col in decimal_cols:
                        if col in df_copy.columns:
                            df_copy[col] = df_copy[col].apply(
                                lambda x: f"{x:.2f}" if pd.notna(x) else ""
                            )

                    # 시가총액 (억 달러)
                    if 'MktCap($B)' in df_copy.columns:
                        df_copy['MktCap($B)'] = df_copy['MktCap($B)'].apply(
                            lambda x: f"{x:.1f}B" if pd.notna(x) else ""
                        )

                    df_copy.to_excel(writer, sheet_name=profile[:30], index=False, startrow=5)
                    ExcelExporter._format_sheet(writer, profile[:30], df_copy, profile=profile)

        print(f"\n📁 결과 저장 완료: {filename}")
        return filename

    @staticmethod
    def _create_column_guide(writer):
        """Column Guide 시트 생성"""
        guide_data = []
        for col, desc in ExcelExporter.COLUMN_DESCRIPTIONS.items():
            guide_data.append({'컬럼명': col, '설명': desc})

        guide_df = pd.DataFrame(guide_data)
        guide_df.to_excel(writer, sheet_name='Column_Guide', index=False)

        # 포맷 적용
        worksheet = writer.sheets['Column_Guide']

        # 헤더 스타일
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)

        for col in range(1, 3):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 컬럼 너비 설정
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 60

        # 행 높이 및 정렬
        for row in range(2, len(guide_data) + 2):
            worksheet.row_dimensions[row].height = 20
            worksheet.cell(row=row, column=1).alignment = Alignment(vertical='center')
            worksheet.cell(row=row, column=2).alignment = Alignment(vertical='center', wrap_text=True)

        # 틀 고정
        worksheet.freeze_panes = 'A2'

    @staticmethod
    def _format_sheet(writer, sheet_name: str, df: pd.DataFrame, is_summary: bool = False, profile: str = None):
        """시트 포맷 적용"""
        worksheet = writer.sheets[sheet_name]

        # 전략별 필터 기준 헤더 추가 (Summary 제외)
        if not is_summary and profile and profile in ExcelExporter.STRATEGY_CRITERIA:
            criteria_info = ExcelExporter.STRATEGY_CRITERIA[profile]

            # 전략 이름 (1행)
            worksheet.merge_cells('A1:C1')
            title_cell = worksheet['A1']
            title_cell.value = f"📊 {criteria_info['name']}"
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            worksheet.row_dimensions[1].height = 25

            # 필터 기준 (2-4행)
            worksheet.merge_cells('A2:C2')
            criteria_cell = worksheet['A2']
            criteria_cell.value = "📋 필터 기준:"
            criteria_cell.font = Font(bold=True, size=11)
            criteria_cell.fill = PatternFill(start_color="D6E4F5", end_color="D6E4F5", fill_type="solid")
            criteria_cell.alignment = Alignment(horizontal='left', vertical='center')

            criteria_text = '\n'.join([f"• {c}" for c in criteria_info['criteria']])
            worksheet.merge_cells('A3:C4')
            criteria_content = worksheet['A3']
            criteria_content.value = criteria_text
            criteria_content.font = Font(size=10)
            criteria_content.fill = PatternFill(start_color="F2F6FC", end_color="F2F6FC", fill_type="solid")
            criteria_content.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            worksheet.row_dimensions[3].height = 30
            worksheet.row_dimensions[4].height = 30

            # 빈 행 (5행)
            worksheet.row_dimensions[5].height = 5

            # 데이터 헤더 행 (6행, startrow=5이므로 실제로는 6행부터)
            header_row = 6

        elif is_summary:
            header_row = 1
        else:
            header_row = 1

        # 헤더 스타일
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 컬럼 너비 자동 조정
        for idx, col in enumerate(df.columns, 1):
            max_length = len(str(col))
            start_row = header_row + 1
            for row in range(start_row, min(start_row + len(df), start_row + 100)):
                try:
                    cell_value = worksheet.cell(row=row, column=idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 35)
            worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width

        # 틀 고정
        if not is_summary and profile:
            worksheet.freeze_panes = f'C{header_row + 1}'
        else:
            worksheet.freeze_panes = 'C2'


# ============================================================================
# 메인 실행 함수
# ============================================================================

def main(csv_file: str = "details_cache_us_all.csv"):
    """메인 실행 함수"""
    try:
        # 스크리너 인스턴스 생성
        screener = StockScreener()

        # 스크리닝 실행
        results = screener.screen_stocks(csv_file)

        if not results:
            print("\n❌ 조건을 충족하는 종목이 없습니다.")
            return None

        # 엑셀 출력
        output_file = ExcelExporter.export(results)

        # 결과 요약 출력
        print("\n" + "=" * 60)
        print("📊 스크리닝 결과 요약")
        print("=" * 60)

        for profile, df in results.items():
            if not df.empty:
                print(f"\n[{profile}]")
                print(f"  • 종목 수: {len(df)}개")
                print(f"  • 평균 점수: {df['TotalScore'].mean():.1f}")
                print(f"  • Top 3: {', '.join(df.head(3)['Ticker'].tolist())}")

                # 추가 통계
                if 'PE' in df.columns:
                    print(f"  • 평균 PE: {df['PE'].mean():.1f}")
                if 'EPS_Growth_3Y' in df.columns:
                    print(f"  • 평균 EPS 성장률: {df['EPS_Growth_3Y'].mean()*100:.1f}%")

        print("\n✅ 스크리닝 완료!")
        return results

    except FileNotFoundError as e:
        print(f"\n❌ 오류: {e}")
        print("CSV 파일을 다운로드하고 경로를 확인해주세요.")
        return None
    except Exception as e:
        print(f"\n❌ 예상치 못한 오류: {e}")
        import traceback
        traceback.print_exc()
        return None


if __name__ == "__main__":
    # CSV 파일 경로를 인자로 전달 가능
    import sys

    csv_path = sys.argv[1] if len(sys.argv) > 1 else "details_cache_us_all.csv"

    print("\n" + "=" * 60)
    print("🚀 개선된 미국 주식 스크리너")
    print("=" * 60)
    print("\n전략:")
    print("  1. undervalued_quality: 저평가 우량주 (워렌 버핏 스타일)")
    print("  2. value_basic: 기본 가치투자")
    print("  3. value_strict: 엄격한 가치투자")
    print("  4. growth_quality: 성장+품질 (장타)")
    print("  5. momentum: 모멘텀 트레이딩 (단타)")
    print("  6. swing: 스윙 트레이딩 (단타)")
    print("=" * 60 + "\n")

    main(csv_path)
